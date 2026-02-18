import streamlit as st
from datetime import datetime, timedelta
import hashlib
import json
import random
import string
from pathlib import Path
from PIL import Image
import base64
from io import BytesIO
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import time
import qrcode
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import requests
import zipfile
import shutil
import csv
import openpyxl
from fpdf import FPDF
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from wordcloud import WordCloud
import networkx as nx
from sklearn import preprocessing
from sklearn.cluster import KMeans
import joblib
import tensorflow as tf
from tensorflow import keras
from tensorflow.keras import layers
import plotly.figure_factory as ff
import calendar
from datetime import date
import uuid
import re
import bcrypt
import jwt
import secrets
import hashlib
import hmac
import pickle
import sqlite3
import mysql.connector
from mysql.connector import Error
import pymongo
from pymongo import MongoClient
import redis
import celery
from celery import Celery
import asyncio
import aiohttp
import websockets
import socketio
import flask
from flask import Flask, request, jsonify
import django
from django.conf import settings
import fastapi
from fastapi import FastAPI, HTTPException
import uvicorn
import gunicorn
import nginx
import apache
import docker
from docker import DockerClient
import kubernetes
from kubernetes import client, config
import awscli
import boto3
from boto3 import session
import azure
from azure.storage.blob import BlobServiceClient
import google.cloud
from google.cloud import storage
import firebase_admin
from firebase_admin import credentials, firestore
import stripe
import paypalrestsdk
import mpesa
from mpesa import MpesaClient
import africastalking
import twilio
from twilio.rest import Client
import vonage
import nexmo
import telegram
from telegram import Bot
import discord
from discord import Webhook
import slack
from slack_sdk import WebClient
import zoom
from zoomus import ZoomClient
import google_meet
import microsoft_teams
from microsoft_teams import TeamsClient
import whatsapp
from whatsapp import WhatsApp
import facebook
from facebook import GraphAPI
import instagram
from instagram import InstagramAPI
import tiktok
from tiktok import TikTokAPI
import youtube
from youtube import YouTubeAPI
import twitter
from twitter import TwitterAPI
import linkedin
from linkedin import LinkedInAPI
import pinterest
from pinterest import PinterestAPI
import snapchat
from snapchat import SnapchatAPI
import reddit
from reddit import RedditAPI
import quora
from quora import QuoraAPI
import medium
from medium import MediumAPI
import substack
from substack import SubstackAPI
import wordpress
from wordpress import WordPressAPI
import wix
from wix import WixAPI
import squarespace
from squarespace import SquarespaceAPI
import shopify
from shopify import ShopifyAPI
import woocommerce
from woocommerce import WooCommerceAPI
import magento
from magento import MagentoAPI
import prestashop
from prestashop import PrestaShopAPI
import opencart
from opencart import OpenCartAPI
import bigcommerce
from bigcommerce import BigCommerceAPI
import salesforce
from salesforce import SalesforceAPI
import hubspot
from hubspot import HubSpotAPI
import zoho
from zoho import ZohoAPI
import sap
from sap import SAPAPI
import oracle
from oracle import OracleAPI
import microsoft_dynamics
from microsoft_dynamics import DynamicsAPI
import odoo
from odoo import OdooAPI
import erpnext
from erpnext import ERPNextAPI
import dolibarr
from dolibarr import DolibarrAPI
import vtiger
from vtiger import VtigerAPI
import suitecrm
from suitecrm import SuiteCRMAPI
import sugarcrm
from sugarcrm import SugarCRMAPI
import cognos
from cognos import CognosAPI
import tableau
from tableau import TableauAPI
import powerbi
from powerbi import PowerBIAPI
import looker
from looker import LookerAPI
import metabase
from metabase import MetabaseAPI
import superset
from superset import SupersetAPI
import redash
from redash import RedashAPI
import grafana
from grafana import GrafanaAPI
import kibana
from kibana import KibanaAPI
import elasticsearch
from elasticsearch import Elasticsearch
import logstash
import beats
from beats import Filebeat, Metricbeat
import splunk
from splunk import SplunkAPI
import datadog
from datadog import DatadogAPI
import newrelic
from newrelic import NewRelicAPI
import dynatrace
from dynatrace import DynatraceAPI
import appdynamics
from appdynamics import AppDynamicsAPI
import prometheus
from prometheus import PrometheusAPI
import influxdb
from influxdb import InfluxDBClient
import grafana
from grafana import GrafanaAPI
import kibana
from kibana import KibanaAPI
import elasticsearch
from elasticsearch import Elasticsearch
import logstash
import beats
from beats import Filebeat, Metricbeat
import splunk
from splunk import SplunkAPI
import datadog
from datadog import DatadogAPI
import newrelic
from newrelic import NewRelicAPI
import dynatrace
from dynatrace import DynatraceAPI
import appdynamics
from appdynamics import AppDynamicsAPI
import prometheus
from prometheus import PrometheusAPI
import influxdb
from influxdb import InfluxDBClient

# ============ PAGE CONFIG ============
st.set_page_config(
    page_title="✨ Complete School Community Hub ✨",
    page_icon="🌟",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============ RESPONSIVE DESIGN META TAG ============
st.markdown("""
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=yes">
<style>
    @media (max-width: 768px) {
        .main .block-container { padding: 0.8rem !important; }
        h1 { font-size: 1.8rem !important; }
        h2 { font-size: 1.3rem !important; }
        .stButton button { font-size: 0.9rem !important; padding: 0.4rem 0.8rem !important; }
    }
    
    /* Print styles */
    @media print {
        .no-print { display: none !important; }
        .print-only { display: block !important; }
        body { background: white; }
        .main .block-container { background: white; box-shadow: none; }
    }
</style>
""", unsafe_allow_html=True)

# ============ KENYAN CURRICULUM DATA ============
PRIMARY_SUBJECTS = [
    "Mathematics", "English", "Kiswahili", "Science and Technology",
    "Social Studies", "CRE / IRE / HRE", "Agriculture", "Home Science",
    "Art and Craft", "Music", "Physical Education"
]

JUNIOR_SECONDARY_SUBJECTS = [
    "Mathematics", "English", "Kiswahili", "Integrated Science",
    "Social Studies", "CRE / IRE / HRE", "Business Studies",
    "Agriculture", "Home Science", "Computer Science",
    "Pre-Technical Studies", "Visual Arts", "Performing Arts",
    "Physical Education"
]

SENIOR_SECONDARY_SUBJECTS = {
    "Mathematics": ["Mathematics"],
    "English": ["English"],
    "Kiswahili": ["Kiswahili"],
    "Sciences": ["Biology", "Chemistry", "Physics", "General Science"],
    "Humanities": ["History", "Geography", "CRE", "IRE", "HRE"],
    "Technical": ["Computer Studies", "Business Studies", "Agriculture", "Home Science"],
    "Languages": ["French", "German", "Arabic", "Sign Language"]
}

KENYAN_GRADES = [
    "Grade 1 (7 subjects)", "Grade 2 (7 subjects)", "Grade 3 (7 subjects)",
    "Grade 4 (7 subjects)", "Grade 5 (7 subjects)", "Grade 6 (7 subjects)",
    "Grade 7 (12 subjects)", "Grade 8 (12 subjects)", "Grade 9 (12 subjects)",
    "Form 1 (11 subjects)", "Form 2 (11 subjects)", "Form 3 (11 subjects)", "Form 4 (11 subjects)"
]

STREAMS = ["East", "West", "North", "South", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]

TERMS = ["Term 1", "Term 2", "Term 3"]

EXAM_TYPES = ["End Term", "Mid Term", "CAT 1", "CAT 2", "CAT 3", "Opener", "Closer", "Joint", "Mock", "KCPE", "KCSE"]

def get_subjects_for_grade(grade):
    if "Grade" in grade and any(str(i) in grade for i in range(1, 7)):
        return PRIMARY_SUBJECTS
    elif "Grade" in grade and any(str(i) in grade for i in range(7, 10)):
        return JUNIOR_SECONDARY_SUBJECTS
    elif "Form" in grade:
        subjects = []
        for category, subj_list in SENIOR_SECONDARY_SUBJECTS.items():
            subjects.extend(subj_list)
        return subjects
    else:
        return PRIMARY_SUBJECTS

# ============ BEAUTIFUL GRADIENT BACKGROUND ============
def get_gradient_colors():
    """Returns a set of beautiful flowing gradient colors"""
    gradients = [
        # Sunrise gradient
        """
        background: linear-gradient(-45deg, 
            #ff6b6b, #feca57, #ff9ff3, #48dbfb, #1dd1a1, #f368e0, #ff9f43
        );
        background-size: 400% 400%;
        animation: gradient 15s ease infinite;
        """,
        # Ocean sunset
        """
        background: linear-gradient(-45deg, 
            #ff0844, #ffb199, #ff0844, #00d2ff, #3a1c71, #d76d77, #ffaf7b
        );
        background-size: 400% 400%;
        animation: gradient 18s ease infinite;
        """,
        # Purple haze
        """
        background: linear-gradient(-45deg, 
            #8E2DE2, #4A00E0, #6a3093, #a044ff, #c471ed, #f64f59, #c471ed
        );
        background-size: 400% 400%;
        animation: gradient 20s ease infinite;
        """,
        # Tropical
        """
        background: linear-gradient(-45deg, 
            #00b09b, #96c93d, #c6ffdd, #fbd786, #f7797d, #4facfe, #00f2fe
        );
        background-size: 400% 400%;
        animation: gradient 16s ease infinite;
        """,
        # Cherry blossom
        """
        background: linear-gradient(-45deg, 
            #ff9a9e, #fad0c4, #fad0c4, #ffd1ff, #a1c4fd, #c2e9fb, #fbc2eb
        );
        background-size: 400% 400%;
        animation: gradient 22s ease infinite;
        """,
        # Midnight city
        """
        background: linear-gradient(-45deg, 
            #232526, #414345, #232526, #2c3e50, #4b6cb7, #182848, #4b6cb7
        );
        background-size: 400% 400%;
        animation: gradient 25s ease infinite;
        """,
        # Autumn leaves
        """
        background: linear-gradient(-45deg, 
            #e44d2e, #f39c12, #d35400, #e67e22, #f1c40f, #e67e22, #d35400
        );
        background-size: 400% 400%;
        animation: gradient 19s ease infinite;
        """,
        # Northern lights
        """
        background: linear-gradient(-45deg, 
            #43C6AC, #191654, #43C6AC, #02AAB0, #00CDAC, #02AAB0, #191654
        );
        background-size: 400% 400%;
        animation: gradient 21s ease infinite;
        """
    ]
    return random.choice(gradients)

# ============ LUMINOUS SECTION BACKGROUNDS ============
def get_luminous_colors():
    """Returns vibrant, glowing background colors for sections"""
    luminous_colors = [
        """
        background: linear-gradient(135deg, #FF6B6B, #FF8E8E);
        box-shadow: 0 0 30px rgba(255, 107, 107, 0.5);
        """,
        """
        background: linear-gradient(135deg, #4ECDC4, #6EE7E7);
        box-shadow: 0 0 30px rgba(78, 205, 196, 0.5);
        """,
        """
        background: linear-gradient(135deg, #45B7D1, #6EC8E0);
        box-shadow: 0 0 30px rgba(69, 183, 209, 0.5);
        """,
        """
        background: linear-gradient(135deg, #96CEB4, #B8E0CC);
        box-shadow: 0 0 30px rgba(150, 206, 180, 0.5);
        """,
        """
        background: linear-gradient(135deg, #FFEEAD, #FFF2C2);
        box-shadow: 0 0 30px rgba(255, 238, 173, 0.5);
        """,
        """
        background: linear-gradient(135deg, #D4A5A5, #E3C0C0);
        box-shadow: 0 0 30px rgba(212, 165, 165, 0.5);
        """,
        """
        background: linear-gradient(135deg, #9B59B6, #B07CC6);
        box-shadow: 0 0 30px rgba(155, 89, 182, 0.5);
        """,
        """
        background: linear-gradient(135deg, #3498DB, #5DADE2);
        box-shadow: 0 0 30px rgba(52, 152, 219, 0.5);
        """,
        """
        background: linear-gradient(135deg, #E67E22, #F39C12);
        box-shadow: 0 0 30px rgba(230, 126, 34, 0.5);
        """,
        """
        background: linear-gradient(135deg, #27AE60, #2ECC71);
        box-shadow: 0 0 30px rgba(39, 174, 96, 0.5);
        """
    ]
    return random.choice(luminous_colors)

# ============ GOLDEN SIDEBAR STYLING ============
def get_golden_gradient():
    """Returns a beautiful golden gradient for sidebar"""
    return """
    background: linear-gradient(135deg, 
        #cfa668, #e5b873, #f5d742, #e6be5a, #d4a545, #c1933a, #ad7e2e
    );
    background-size: 300% 300%;
    animation: golden-shimmer 8s ease infinite;
    """

# ============ CUSTOM CSS ============
GRADIENT_STYLE = get_gradient_colors()

st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700;800&display=swap');
    
    * {{
        font-family: 'Poppins', sans-serif;
        box-sizing: border-box;
    }}
    
    @keyframes gradient {{
        0% {{ background-position: 0% 50%; }}
        50% {{ background-position: 100% 50%; }}
        100% {{ background-position: 0% 50%; }}
    }}
    
    @keyframes golden-shimmer {{
        0% {{ background-position: 0% 50%; }}
        50% {{ background-position: 100% 50%; }}
        100% {{ background-position: 0% 50%; }}
    }}
    
    /* Apply gradient to body background */
    body {{
        {GRADIENT_STYLE}
        margin: 0;
        padding: 0;
        min-height: 100vh;
    }}
    
    .stApp {{
        background: transparent !important;
    }}
    
    /* Main content area */
    .main .block-container {{
        background: transparent !important;
        padding: 1rem;
        margin: 0;
        border: none;
        box-shadow: none;
    }}
    
    /* Section containers with luminous backgrounds */
    .section-container {{
        border-radius: 20px;
        padding: 2rem;
        margin: 1rem 0;
        animation: glow 3s ease-in-out infinite;
    }}
    
    @keyframes glow {{
        0% {{ box-shadow: 0 0 20px rgba(255, 255, 255, 0.3); }}
        50% {{ box-shadow: 0 0 40px rgba(255, 255, 255, 0.6); }}
        100% {{ box-shadow: 0 0 20px rgba(255, 255, 255, 0.3); }}
    }}
    
    .section-community {{ {get_luminous_colors()} }}
    .section-management {{ {get_luminous_colors()} }}
    .section-personal {{ {get_luminous_colors()} }}
    
    /* Golden Sidebar */
    section[data-testid="stSidebar"] {{
        {get_golden_gradient()}
        backdrop-filter: blur(5px) !important;
        border-right: 2px solid rgba(255, 215, 0, 0.4) !important;
        box-shadow: 5px 0 30px rgba(218, 165, 32, 0.4) !important;
        z-index: 20;
    }}
    
    section[data-testid="stSidebar"] > div {{
        background: rgba(0, 0, 0, 0.2) !important;
        padding: 1rem 0.8rem !important;
        width: 100% !important;
    }}
    
    section[data-testid="stSidebar"] .stMarkdown,
    section[data-testid="stSidebar"] .stRadio label,
    section[data-testid="stSidebar"] p,
    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3,
    section[data-testid="stSidebar"] span,
    section[data-testid="stSidebar"] div {{
        color: #FFD700 !important;
        text-shadow: 1px 1px 2px rgba(0,0,0,0.5) !important;
        font-weight: 600 !important;
    }}
    
    section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] {{
        background: rgba(0, 0, 0, 0.3) !important;
        border-radius: 12px !important;
        padding: 0.5rem !important;
        border: 1px solid rgba(255, 215, 0, 0.3) !important;
        margin: 0.8rem 0 !important;
    }}
    
    section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label {{
        background: transparent !important;
        border-radius: 8px !important;
        padding: 8px 10px !important;
        margin: 2px 0 !important;
        transition: all 0.2s ease !important;
        color: #FFD700 !important;
        font-weight: 600 !important;
        text-shadow: 1px 1px 2px rgba(0,0,0,0.5) !important;
    }}
    
    section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label:hover {{
        background: rgba(255, 215, 0, 0.2) !important;
        transform: translateX(5px) !important;
    }}
    
    section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label[data-checked="true"] {{
        background: rgba(255, 215, 0, 0.3) !important;
        border-left: 4px solid #FFD700 !important;
        font-weight: 700 !important;
        box-shadow: 0 2px 10px rgba(255, 215, 0, 0.3) !important;
    }}
    
    section[data-testid="stSidebar"] .stButton button {{
        background: linear-gradient(135deg, #FFD700, #DAA520) !important;
        color: #2b2b2b !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 8px 12px !important;
        font-weight: 700 !important;
        font-size: 0.9rem !important;
        transition: all 0.2s ease !important;
        width: 100% !important;
        margin: 0.5rem 0 !important;
        box-shadow: 0 4px 15px rgba(218, 165, 32, 0.4) !important;
    }}
    
    section[data-testid="stSidebar"] .stButton button:hover {{
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(255, 215, 0, 0.6) !important;
    }}
    
    /* School header in sidebar */
    .school-header {{
        background: rgba(0, 0, 0, 0.3);
        border: 1px solid #FFD700;
        border-radius: 12px;
        padding: 12px;
        margin-bottom: 12px;
        text-align: center;
    }}
    
    .school-header h2 {{
        color: #FFD700 !important;
        margin: 0;
        font-size: 1.3rem;
        font-weight: 700;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.5);
    }}
    
    .school-code {{
        background: rgba(0, 0, 0, 0.3);
        padding: 4px;
        border-radius: 20px;
        margin-top: 5px;
        border: 1px solid #FFD700;
    }}
    
    .school-code code {{
        background: transparent !important;
        color: #FFD700 !important;
        font-size: 0.8rem;
        font-weight: 700;
    }}
    
    /* Profile card in sidebar */
    .profile-card {{
        background: rgba(0, 0, 0, 0.3);
        border: 1px solid #FFD700;
        border-radius: 12px;
        padding: 10px;
        margin-bottom: 12px;
        display: flex;
        align-items: center;
        gap: 10px;
    }}
    
    /* Input field styling - FIXED DROPDOWNS WITH CLEAR TEXT */
    .stSelectbox div[data-baseweb="select"] {{
        background: white !important;
        border: 2px solid #FFD700 !important;
        border-radius: 10px !important;
        color: #000000 !important;
        font-weight: 500 !important;
        transition: all 0.3s ease !important;
    }}
    
    .stSelectbox div[data-baseweb="select"]:hover {{
        border-color: #DAA520 !important;
        box-shadow: 0 0 15px rgba(218, 165, 32, 0.3) !important;
    }}
    
    .stSelectbox div[data-baseweb="select"] span {{
        color: #000000 !important;
        font-weight: 500 !important;
    }}
    
    .stSelectbox div[role="listbox"] {{
        background: white !important;
        border: 2px solid #FFD700 !important;
        border-radius: 10px !important;
        box-shadow: 0 5px 20px rgba(0,0,0,0.2) !important;
    }}
    
    .stSelectbox div[role="listbox"] div {{
        color: #000000 !important;
        font-weight: 500 !important;
        padding: 8px 12px !important;
    }}
    
    .stSelectbox div[role="listbox"] div:hover {{
        background: linear-gradient(135deg, #FFD70020, #DAA52020) !important;
    }}
    
    .stTextInput input, 
    .stTextArea textarea, 
    .stDateInput input,
    .stNumberInput input {{
        background: white !important;
        border: 2px solid #FFD700 !important;
        border-radius: 10px !important;
        padding: 0.6rem 1rem !important;
        font-size: 0.95rem !important;
        color: #000000 !important;
        font-weight: 500 !important;
        transition: all 0.3s ease !important;
    }}
    
    .stTextInput input:focus, 
    .stTextArea textarea:focus,
    .stDateInput input:focus,
    .stNumberInput input:focus {{
        border-color: #DAA520 !important;
        box-shadow: 0 0 15px rgba(218, 165, 32, 0.3) !important;
    }}
    
    .stTextInput label,
    .stTextArea label,
    .stSelectbox label,
    .stDateInput label,
    .stNumberInput label {{
        color: #333333 !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
    }}
    
    /* Tabs styling */
    .stTabs [data-baseweb="tab-list"] {{
        background: linear-gradient(135deg, #FFD700, #DAA520) !important;
        border-radius: 12px !important;
        padding: 0.3rem !important;
        gap: 0.3rem;
        margin-bottom: 1.5rem !important;
    }}
    
    .stTabs [data-baseweb="tab"] {{
        color: #2b2b2b !important;
        border-radius: 8px !important;
        padding: 0.5rem 1rem !important;
        font-weight: 600 !important;
    }}
    
    .stTabs [aria-selected="true"] {{
        background: rgba(0, 0, 0, 0.2) !important;
        color: #000000 !important;
    }}
    
    /* Headers */
    h1 {{
        background: linear-gradient(135deg, #FFD700, #FFA500, #FF8C00);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        font-size: 2.5rem !important;
        font-weight: 700 !important;
        text-align: center;
        margin-bottom: 1.5rem !important;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.1);
    }}
    
    /* Golden cards */
    .golden-card {{
        background: rgba(255, 255, 255, 0.9);
        border-left: 6px solid #FFD700;
        border-radius: 12px;
        padding: 20px;
        margin-bottom: 15px;
        box-shadow: 0 4px 15px rgba(218, 165, 32, 0.2);
    }}
    
    /* Performance badges */
    .performance-excellent {{
        background: linear-gradient(135deg, #28a745, #20c997);
        color: white;
        padding: 4px 12px;
        border-radius: 20px;
        font-weight: 600;
        font-size: 0.8rem;
        display: inline-block;
        box-shadow: 0 0 15px rgba(40, 167, 69, 0.5);
    }}
    
    .performance-good {{
        background: linear-gradient(135deg, #17a2b8, #6f42c1);
        color: white;
        padding: 4px 12px;
        border-radius: 20px;
        font-weight: 600;
        font-size: 0.8rem;
        display: inline-block;
        box-shadow: 0 0 15px rgba(23, 162, 184, 0.5);
    }}
    
    .performance-average {{
        background: linear-gradient(135deg, #ffc107, #fd7e14);
        color: white;
        padding: 4px 12px;
        border-radius: 20px;
        font-weight: 600;
        font-size: 0.8rem;
        display: inline-block;
        box-shadow: 0 0 15px rgba(255, 193, 7, 0.5);
    }}
    
    .performance-needs-improvement {{
        background: linear-gradient(135deg, #dc3545, #c82333);
        color: white;
        padding: 4px 12px;
        border-radius: 20px;
        font-weight: 600;
        font-size: 0.8rem;
        display: inline-block;
        box-shadow: 0 0 15px rgba(220, 53, 69, 0.5);
    }}
    
    /* Chat styling */
    .chat-container {{
        background: rgba(255, 255, 255, 0.9);
        border-radius: 16px;
        padding: 20px;
        height: 400px;
        overflow-y: auto;
        display: flex;
        flex-direction: column;
        gap: 15px;
        border: 1px solid #FFD700;
        box-shadow: 0 0 30px rgba(255, 215, 0, 0.3);
    }}
    
    .chat-message-wrapper {{
        display: flex;
        margin-bottom: 10px;
        animation: fadeIn 0.3s ease;
    }}
    
    @keyframes fadeIn {{
        from {{ opacity: 0; transform: translateY(10px); }}
        to {{ opacity: 1; transform: translateY(0); }}
    }}
    
    .chat-message-sent {{
        justify-content: flex-end;
    }}
    
    .chat-message-received {{
        justify-content: flex-start;
    }}
    
    .chat-bubble {{
        max-width: 70%;
        padding: 12px 16px;
        border-radius: 20px;
        position: relative;
        word-wrap: break-word;
    }}
    
    .chat-bubble-sent {{
        background: linear-gradient(135deg, #FFD700, #DAA520);
        color: #2b2b2b;
        border-bottom-right-radius: 4px;
        box-shadow: 0 0 20px rgba(255, 215, 0, 0.5);
    }}
    
    .chat-bubble-received {{
        background: rgba(255, 255, 255, 0.95);
        color: #333333;
        border-bottom-left-radius: 4px;
        border: 1px solid #FFD700;
        box-shadow: 0 0 20px rgba(255, 215, 0, 0.3);
    }}
    
    .chat-sender-info {{
        display: flex;
        align-items: center;
        gap: 8px;
        margin-bottom: 4px;
    }}
    
    .chat-sender-name {{
        font-size: 0.8rem;
        color: #DAA520;
        font-weight: 600;
    }}
    
    .chat-time {{
        font-size: 0.65rem;
        color: rgba(51, 51, 51, 0.5);
        margin-top: 4px;
        text-align: right;
    }}
    
    /* Main navigation buttons on welcome page */
    .nav-button {{
        background: linear-gradient(135deg, #FFD700, #DAA520);
        color: #2b2b2b;
        border: none;
        border-radius: 15px;
        padding: 25px;
        font-size: 1.3rem;
        font-weight: 700;
        text-align: center;
        cursor: pointer;
        transition: all 0.3s ease;
        box-shadow: 0 10px 30px rgba(218, 165, 32, 0.3);
        margin: 10px 0;
    }}
    
    .nav-button:hover {{
        transform: translateY(-5px);
        box-shadow: 0 15px 40px rgba(255, 215, 0, 0.4);
    }}
    
    /* Class cards */
    .class-card {{
        background: rgba(255, 255, 255, 0.95);
        border-radius: 12px;
        padding: 15px;
        margin-bottom: 10px;
        border-left: 4px solid #FFD700;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }}
    
    /* Member cards */
    .member-card {{
        background: rgba(255, 255, 255, 0.95);
        border-radius: 12px;
        padding: 15px;
        margin-bottom: 10px;
        display: flex;
        align-items: center;
        gap: 15px;
        border: 1px solid #FFD700;
        box-shadow: 0 0 20px rgba(255, 215, 0, 0.2);
    }}
    
    .member-pic {{
        width: 50px;
        height: 50px;
        border-radius: 50%;
        object-fit: cover;
        border: 2px solid #FFD700;
    }}
    
    /* Badges */
    .request-badge {{
        background: #FFD700;
        color: #2b2b2b;
        padding: 2px 8px;
        border-radius: 12px;
        font-size: 0.7rem;
        font-weight: 600;
        margin-left: 8px;
        box-shadow: 0 0 10px rgba(255, 215, 0, 0.5);
    }}
    
    /* Metric styling */
    .stMetric {{
        background: rgba(255, 255, 255, 0.9) !important;
        border-radius: 12px !important;
        padding: 1rem !important;
        box-shadow: 0 0 20px rgba(255, 215, 0, 0.3) !important;
    }}
    
    .stMetric label {{
        color: #333333 !important;
        font-weight: 600 !important;
    }}
    
    .stMetric div {{
        color: #FFD700 !important;
        font-weight: 700 !important;
    }}
    
    /* QR Code styling */
    .qr-container {{
        display: flex;
        flex-wrap: wrap;
        gap: 20px;
        justify-content: center;
        margin: 20px 0;
    }}
    
    .qr-item {{
        background: white;
        padding: 15px;
        border-radius: 10px;
        text-align: center;
        box-shadow: 0 0 15px rgba(255, 215, 0, 0.3);
        border: 2px solid #FFD700;
    }}
    
    /* Table styles */
    table {{
        width: 100%;
        border-collapse: collapse;
        margin: 15px 0;
    }}
    
    th {{
        background: linear-gradient(135deg, #FFD700, #DAA520);
        color: #2b2b2b;
        padding: 10px;
        text-align: left;
    }}
    
    td {{
        padding: 8px 10px;
        border-bottom: 1px solid #FFD700;
    }}
    
    tr:hover {{
        background: rgba(255, 215, 0, 0.1);
    }}
    
    /* Print styles */
    .print-only {{
        display: none;
    }}
    
    @media print {{
        .no-print {{
            display: none !important;
        }}
        .print-only {{
            display: block !important;
        }}
        body {{
            background: white;
        }}
        .main .block-container {{
            background: white;
            box-shadow: none;
            padding: 0;
        }}
    }}
</style>
""", unsafe_allow_html=True)

# ============ CODE GENERATOR ============
def generate_id(prefix, length=8):
    chars = string.ascii_uppercase + string.digits
    random_part = ''.join(random.choices(chars, k=length))
    return f"{prefix}{random_part}"

def generate_school_code():
    chars = string.ascii_uppercase + string.digits
    return 'SCH' + ''.join(random.choices(chars, k=6))

def generate_class_code():
    return 'CLS' + ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

def generate_group_code():
    return 'GRP' + ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

def generate_admission_number():
    year = datetime.now().strftime("%y")
    random_num = ''.join(random.choices(string.digits, k=4))
    return f"ADM/{year}/{random_num}"

def generate_teacher_code():
    return 'TCH' + ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

def generate_furniture_code():
    return 'FUR' + ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

def generate_qr_code(data):
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    return img

# ============ DATA STORAGE ============
DATA_DIR = Path("complete_school_data")
DATA_DIR.mkdir(exist_ok=True)

SCHOOLS_FILE = DATA_DIR / "all_schools.json"
USERS_FILE = DATA_DIR / "users.json"
CLASSES_FILE = DATA_DIR / "classes.json"
GROUPS_FILE = DATA_DIR / "groups.json"
ANNOUNCEMENTS_FILE = DATA_DIR / "announcements.json"
ASSIGNMENTS_FILE = DATA_DIR / "assignments.json"
MESSAGES_FILE = DATA_DIR / "messages.json"
FRIEND_REQUESTS_FILE = DATA_DIR / "friend_requests.json"
FRIENDSHIPS_FILE = DATA_DIR / "friendships.json"
GROUP_CHATS_FILE = DATA_DIR / "group_chats.json"
ACADEMIC_RECORDS_FILE = DATA_DIR / "academic_records.json"
ATTENDANCE_FILE = DATA_DIR / "attendance.json"
FEES_FILE = DATA_DIR / "fees.json"
DISCIPLINE_FILE = DATA_DIR / "discipline.json"
TEACHER_REVIEWS_FILE = DATA_DIR / "teacher_reviews.json"
PARENT_FEEDBACK_FILE = DATA_DIR / "parent_feedback.json"
CLASS_REQUESTS_FILE = DATA_DIR / "class_requests.json"
GROUP_REQUESTS_FILE = DATA_DIR / "group_requests.json"
TEACHER_ALLOCATIONS_FILE = DATA_DIR / "teacher_allocations.json"
FURNITURE_ALLOCATIONS_FILE = DATA_DIR / "furniture_allocations.json"
BOOK_CATALOG_FILE = DATA_DIR / "book_catalog.json"
EXAM_RESULTS_FILE = DATA_DIR / "exam_results.json"
TIMETABLE_FILE = DATA_DIR / "timetable.json"
LIBRARY_BOOKS_FILE = DATA_DIR / "library_books.json"
SPORTS_DATA_FILE = DATA_DIR / "sports_data.json"
CLUBS_DATA_FILE = DATA_DIR / "clubs_data.json"
TRANSPORT_DATA_FILE = DATA_DIR / "transport_data.json"
CAFETERIA_DATA_FILE = DATA_DIR / "cafeteria_data.json"
HEALTH_RECORDS_FILE = DATA_DIR / "health_records.json"
COUNSELING_RECORDS_FILE = DATA_DIR / "counseling_records.json"
ALUMNI_DATA_FILE = DATA_DIR / "alumni_data.json"
STAFF_DATA_FILE = DATA_DIR / "staff_data.json"
PARENT_TEACHER_MEETINGS_FILE = DATA_DIR / "parent_teacher_meetings.json"
EXTRACURRICULAR_DATA_FILE = DATA_DIR / "extracurricular_data.json"
SCHOOL_CALENDAR_FILE = DATA_DIR / "school_calendar.json"
NEWSLETTER_FILE = DATA_DIR / "newsletter.json"
BUDGET_FILE = DATA_DIR / "budget.json"
PROCUREMENT_FILE = DATA_DIR / "procurement.json"
ASSETS_FILE = DATA_DIR / "assets.json"
INVENTORY_FILE = DATA_DIR / "inventory.json"
MAINTENANCE_FILE = DATA_DIR / "maintenance.json"
SECURITY_FILE = DATA_DIR / "security.json"
EMERGENCY_CONTACTS_FILE = DATA_DIR / "emergency_contacts.json"
ACCIDENT_REPORTS_FILE = DATA_DIR / "accident_reports.json"
FIRST_AID_FILE = DATA_DIR / "first_aid.json"
INSURANCE_FILE = DATA_DIR / "insurance.json"
SPONSORSHIP_FILE = DATA_DIR / "sponsorship.json"
SCHOLARSHIP_FILE = DATA_DIR / "scholarship.json"
BURSARY_FILE = DATA_DIR / "bursary.json"
DONATIONS_FILE = DATA_DIR / "donations.json"
GRANTS_FILE = DATA_DIR / "grants.json"
PARTNERSHIPS_FILE = DATA_DIR / "partnerships.json"
MOUS_FILE = DATA_DIR / "mous.json"
VISITORS_FILE = DATA_DIR / "visitors.json"
EVENTS_FILE = DATA_DIR / "events.json"
TRIPS_FILE = DATA_DIR / "trips.json"
COMPETITIONS_FILE = DATA_DIR / "competitions.json"
AWARDS_FILE = DATA_DIR / "awards.json"
CERTIFICATES_FILE = DATA_DIR / "certificates.json"
TRANSCRIPTS_FILE = DATA_DIR / "transcripts.json"
REPORTS_FILE = DATA_DIR / "reports.json"
ANALYTICS_FILE = DATA_DIR / "analytics.json"
PREDICTIONS_FILE = DATA_DIR / "predictions.json"
RECOMMENDATIONS_FILE = DATA_DIR / "recommendations.json"
FEEDBACK_FILE = DATA_DIR / "feedback.json"
SURVEYS_FILE = DATA_DIR / "surveys.json"
POLLS_FILE = DATA_DIR / "polls.json"
VOTING_FILE = DATA_DIR / "voting.json"
ELECTIONS_FILE = DATA_DIR / "elections.json"
NOMINATIONS_FILE = DATA_DIR / "nominations.json"
APPLICATIONS_FILE = DATA_DIR / "applications.json"
INTERVIEWS_FILE = DATA_DIR / "interviews.json"
RECRUITMENT_FILE = DATA_DIR / "recruitment.json"
TRAINING_FILE = DATA_DIR / "training.json"
WORKSHOPS_FILE = DATA_DIR / "workshops.json"
SEMINARS_FILE = DATA_DIR / "seminars.json"
CONFERENCES_FILE = DATA_DIR / "conferences.json"
WEBINARS_FILE = DATA_DIR / "webinars.json"
ONLINE_COURSES_FILE = DATA_DIR / "online_courses.json"
E_LEARNING_FILE = DATA_DIR / "e_learning.json"
DIGITAL_LIBRARY_FILE = DATA_DIR / "digital_library.json"
RESEARCH_FILE = DATA_DIR / "research.json"
PUBLICATIONS_FILE = DATA_DIR / "publications.json"
JOURNALS_FILE = DATA_DIR / "journals.json"
ARTICLES_FILE = DATA_DIR / "articles.json"
BLOGS_FILE = DATA_DIR / "blogs.json"
FORUMS_FILE = DATA_DIR / "forums.json"
DISCUSSIONS_FILE = DATA_DIR / "discussions.json"
Q_A_FILE = DATA_DIR / "q_a.json"
HELP_DESK_FILE = DATA_DIR / "help_desk.json"
SUPPORT_TICKETS_FILE = DATA_DIR / "support_tickets.json"
TECH_SUPPORT_FILE = DATA_DIR / "tech_support.json"
IT_ASSETS_FILE = DATA_DIR / "it_assets.json"
SOFTWARE_LICENSES_FILE = DATA_DIR / "software_licenses.json"
HARDWARE_FILE = DATA_DIR / "hardware.json"
NETWORK_FILE = DATA_DIR / "network.json"
SERVERS_FILE = DATA_DIR / "servers.json"
BACKUPS_FILE = DATA_DIR / "backups.json"
DISASTER_RECOVERY_FILE = DATA_DIR / "disaster_recovery.json"
BUSINESS_CONTINUITY_FILE = DATA_DIR / "business_continuity.json"
RISK_MANAGEMENT_FILE = DATA_DIR / "risk_management.json"
COMPLIANCE_FILE = DATA_DIR / "compliance.json"
AUDIT_FILE = DATA_DIR / "audit.json"
QUALITY_ASSURANCE_FILE = DATA_DIR / "quality_assurance.json"
PERFORMANCE_REVIEWS_FILE = DATA_DIR / "performance_reviews.json"
APPRAISALS_FILE = DATA_DIR / "appraisals.json"
PROMOTIONS_FILE = DATA_DIR / "promotions.json"
TRANSFERS_FILE = DATA_DIR / "transfers.json"
RETIREMENTS_FILE = DATA_DIR / "retirements.json"
RESIGNATIONS_FILE = DATA_DIR / "resignations.json"
TERMINATIONS_FILE = DATA_DIR / "terminations.json"
LEAVE_REQUESTS_FILE = DATA_DIR / "leave_requests.json"
HOLIDAYS_FILE = DATA_DIR / "holidays.json"
ATTENDANCE_REQUESTS_FILE = DATA_DIR / "attendance_requests.json"
OVERTIME_FILE = DATA_DIR / "overtime.json"
PAYROLL_FILE = DATA_DIR / "payroll.json"
SALARIES_FILE = DATA_DIR / "salaries.json"
ALLOWANCES_FILE = DATA_DIR / "allowances.json"
DEDUCTIONS_FILE = DATA_DIR / "deductions.json"
BENEFITS_FILE = DATA_DIR / "benefits.json"
TAXES_FILE = DATA_DIR / "taxes.json"
PENSIONS_FILE = DATA_DIR / "pensions.json"
LOANS_FILE = DATA_DIR / "loans.json"
ADVANCES_FILE = DATA_DIR / "advances.json"
CLAIMS_FILE = DATA_DIR / "claims.json"
REIMBURSEMENTS_FILE = DATA_DIR / "reimbursements.json"
PETTY_CASH_FILE = DATA_DIR / "petty_cash.json"
BANK_ACCOUNTS_FILE = DATA_DIR / "bank_accounts.json"
CASH_FLOW_FILE = DATA_DIR / "cash_flow.json"
INVESTMENTS_FILE = DATA_DIR / "investments.json"
ASSETS_LIABILITIES_FILE = DATA_DIR / "assets_liabilities.json"
EQUITY_FILE = DATA_DIR / "equity.json"
REVENUE_FILE = DATA_DIR / "revenue.json"
EXPENSES_FILE = DATA_DIR / "expenses.json"
PROFIT_LOSS_FILE = DATA_DIR / "profit_loss.json"
BALANCE_SHEET_FILE = DATA_DIR / "balance_sheet.json"
FINANCIAL_RATIOS_FILE = DATA_DIR / "financial_ratios.json"
BUDGET_VARIANCE_FILE = DATA_DIR / "budget_variance.json"
FORECASTS_FILE = DATA_DIR / "forecasts.json"
PROJECTIONS_FILE = DATA_DIR / "projections.json"
SCENARIOS_FILE = DATA_DIR / "scenarios.json"
STRATEGIC_PLAN_FILE = DATA_DIR / "strategic_plan.json"
VISION_MISSION_FILE = DATA_DIR / "vision_mission.json"
CORE_VALUES_FILE = DATA_DIR / "core_values.json"
SWOT_ANALYSIS_FILE = DATA_DIR / "swot_analysis.json"
PESTLE_ANALYSIS_FILE = DATA_DIR / "pestle_analysis.json"
COMPETITOR_ANALYSIS_FILE = DATA_DIR / "competitor_analysis.json"
MARKET_RESEARCH_FILE = DATA_DIR / "market_research.json"
BRANDING_FILE = DATA_DIR / "branding.json"
MARKETING_FILE = DATA_DIR / "marketing.json"
SOCIAL_MEDIA_FILE = DATA_DIR / "social_media.json"
WEBSITE_FILE = DATA_DIR / "website.json"
NEWSLETTERS_FILE = DATA_DIR / "newsletters.json"
PRESS_RELEASES_FILE = DATA_DIR / "press_releases.json"
MEDIA_COVERAGE_FILE = DATA_DIR / "media_coverage.json"
PUBLIC_RELATIONS_FILE = DATA_DIR / "public_relations.json"
COMMUNITY_OUTREACH_FILE = DATA_DIR / "community_outreach.json"
CORPORATE_SOCIAL_RESPONSIBILITY_FILE = DATA_DIR / "corporate_social_responsibility.json"
SUSTAINABILITY_FILE = DATA_DIR / "sustainability.json"
ENVIRONMENTAL_IMPACT_FILE = DATA_DIR / "environmental_impact.json"
CARBON_FOOTPRINT_FILE = DATA_DIR / "carbon_footprint.json"
GREEN_INITIATIVES_FILE = DATA_DIR / "green_initiatives.json"
WASTE_MANAGEMENT_FILE = DATA_DIR / "waste_management.json"
RECYCLING_FILE = DATA_DIR / "recycling.json"
ENERGY_EFFICIENCY_FILE = DATA_DIR / "energy_efficiency.json"
WATER_CONSERVATION_FILE = DATA_DIR / "water_conservation.json"
RENEWABLE_ENERGY_FILE = DATA_DIR / "renewable_energy.json"

def load_all_schools():
    if SCHOOLS_FILE.exists():
        with open(SCHOOLS_FILE, 'r') as f:
            return json.load(f)
    return {}

def save_all_schools(schools):
    with open(SCHOOLS_FILE, 'w') as f:
        json.dump(schools, f, indent=2)

def load_school_data(school_code, filename, default):
    if not school_code:
        return default
    filepath = DATA_DIR / f"{school_code}_{filename}"
    if filepath.exists():
        try:
            with open(filepath, 'r') as f:
                return json.load(f)
        except:
            return default
    return default

def save_school_data(school_code, filename, data):
    if school_code:
        with open(DATA_DIR / f"{school_code}_{filename}", 'w') as f:
            json.dump(data, f, indent=2)

# ============ ENCRYPTION FUNCTIONS ============
SECRET_KEY = secrets.token_bytes(32)

def hash_password(password):
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode(), salt).decode()

def verify_password(password, hashed):
    return bcrypt.checkpw(password.encode(), hashed.encode())

def encrypt_data(data):
    # Simple encryption for demo - in production use proper encryption
    return base64.b64encode(json.dumps(data).encode()).decode()

def decrypt_data(data):
    return json.loads(base64.b64decode(data).decode())

def generate_jwt_token(user_id, role):
    payload = {
        'user_id': user_id,
        'role': role,
        'exp': datetime.utcnow() + timedelta(days=1)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm='HS256')

def verify_jwt_token(token):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None

# ============ CHAT & FRIENDSHIP FUNCTIONS ============
def send_friend_request(school_code, from_email, to_email):
    requests = load_school_data(school_code, "friend_requests.json", [])
    if not any(r['from'] == from_email and r['to'] == to_email and r['status'] == 'pending' for r in requests):
        requests.append({
            "id": generate_id("FRQ"),
            "from": from_email,
            "to": to_email,
            "status": "pending",
            "date": datetime.now().strftime("%Y-%m-%d %H:%M")
        })
        save_school_data(school_code, "friend_requests.json", requests)
        return True
    return False

def accept_friend_request(school_code, request_id):
    requests = load_school_data(school_code, "friend_requests.json", [])
    friendships = load_school_data(school_code, "friendships.json", [])
    
    for req in requests:
        if req['id'] == request_id:
            req['status'] = 'accepted'
            friendships.append({
                "user1": min(req['from'], req['to']),
                "user2": max(req['from'], req['to']),
                "since": datetime.now().strftime("%Y-%m-%d %H:%M")
            })
            break
    
    save_school_data(school_code, "friend_requests.json", requests)
    save_school_data(school_code, "friendships.json", friendships)

def decline_friend_request(school_code, request_id):
    requests = load_school_data(school_code, "friend_requests.json", [])
    for req in requests:
        if req['id'] == request_id:
            req['status'] = 'declined'
            break
    save_school_data(school_code, "friend_requests.json", requests)

def get_friends(school_code, user_email):
    friendships = load_school_data(school_code, "friendships.json", [])
    friends = []
    for f in friendships:
        if f['user1'] == user_email:
            friends.append(f['user2'])
        elif f['user2'] == user_email:
            friends.append(f['user1'])
    return friends

def get_pending_requests(school_code, user_email):
    requests = load_school_data(school_code, "friend_requests.json", [])
    return [r for r in requests if r['to'] == user_email and r['status'] == 'pending']

def get_sent_requests(school_code, user_email):
    requests = load_school_data(school_code, "friend_requests.json", [])
    return [r for r in requests if r['from'] == user_email and r['status'] == 'pending']

def send_message(school_code, sender_email, recipient_email, message, attachment=None):
    messages = load_school_data(school_code, "messages.json", [])
    messages.append({
        "id": generate_id("MSG"),
        "sender": sender_email,
        "recipient": recipient_email,
        "message": message,
        "attachment": attachment,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "read": False,
        "deleted": False,
        "conversation_id": f"{min(sender_email, recipient_email)}_{max(sender_email, recipient_email)}"
    })
    save_school_data(school_code, "messages.json", messages)

def mark_as_read(message_id, school_code):
    messages = load_school_data(school_code, "messages.json", [])
    for msg in messages:
        if msg['id'] == message_id:
            msg['read'] = True
            break
    save_school_data(school_code, "messages.json", messages)

def get_unread_count(user_email, school_code):
    messages = load_school_data(school_code, "messages.json", [])
    return len([m for m in messages if m['recipient'] == user_email and not m.get('read', False) and not m.get('deleted', False)])

def get_conversations(user_email, school_code):
    messages = load_school_data(school_code, "messages.json", [])
    conversations = {}
    for msg in messages:
        if not msg.get('deleted', False) and (msg['sender'] == user_email or msg['recipient'] == user_email):
            other = msg['recipient'] if msg['sender'] == user_email else msg['sender']
            if other not in conversations:
                conversations[other] = []
            conversations[other].append(msg)
    for conv in conversations:
        conversations[conv].sort(key=lambda x: x['timestamp'])
    return conversations

# ============ GROUP CHAT FUNCTIONS ============
def create_group_chat(school_code, group_name, created_by, members):
    group_chats = load_school_data(school_code, "group_chats.json", [])
    group_chat = {
        "id": generate_id("GPC"),
        "name": group_name,
        "created_by": created_by,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "members": members,
        "messages": [],
        "admins": [created_by]
    }
    group_chats.append(group_chat)
    save_school_data(school_code, "group_chats.json", group_chats)
    return group_chat['id']

def send_group_message(school_code, group_id, sender_email, message, attachment=None):
    group_chats = load_school_data(school_code, "group_chats.json", [])
    for group in group_chats:
        if group['id'] == group_id:
            group['messages'].append({
                "id": generate_id("GPM"),
                "sender": sender_email,
                "message": message,
                "attachment": attachment,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "read_by": [sender_email]
            })
            break
    save_school_data(school_code, "group_chats.json", group_chats)

def get_user_groups(school_code, user_email):
    groups = load_school_data(school_code, "groups.json", [])
    group_chats = load_school_data(school_code, "group_chats.json", [])
    user_groups = []
    
    for group in groups:
        if user_email in group.get('members', []):
            user_groups.append({
                "id": group['code'],
                "name": group['name'],
                "type": "regular",
                "members": group.get('members', []),
                "chat_id": None
            })
    
    for chat in group_chats:
        if user_email in chat.get('members', []):
            user_groups.append({
                "id": chat['id'],
                "name": chat['name'],
                "type": "chat",
                "members": chat.get('members', []),
                "chat_id": chat['id']
            })
    
    return user_groups

def get_group_members(school_code, group_id):
    group_chats = load_school_data(school_code, "group_chats.json", [])
    for group in group_chats:
        if group['id'] == group_id:
            return group.get('members', [])
    return []

def add_group_admin(school_code, group_id, user_email):
    group_chats = load_school_data(school_code, "group_chats.json", [])
    for group in group_chats:
        if group['id'] == group_id:
            if 'admins' not in group:
                group['admins'] = []
            if user_email not in group['admins']:
                group['admins'].append(user_email)
            break
    save_school_data(school_code, "group_chats.json", group_chats)

def remove_group_member(school_code, group_id, user_email):
    group_chats = load_school_data(school_code, "group_chats.json", [])
    for group in group_chats:
        if group['id'] == group_id:
            if user_email in group.get('members', []):
                group['members'].remove(user_email)
            if user_email in group.get('admins', []):
                group['admins'].remove(user_email)
            break
    save_school_data(school_code, "group_chats.json", group_chats)

# ============ ATTACHMENT FUNCTIONS ============
def save_attachment(uploaded_file):
    if uploaded_file:
        bytes_data = uploaded_file.getvalue()
        b64 = base64.b64encode(bytes_data).decode()
        return {
            "name": uploaded_file.name,
            "type": uploaded_file.type,
            "data": b64,
            "size": len(bytes_data)
        }
    return None

def display_attachment(attachment):
    if attachment:
        file_ext = attachment['name'].split('.')[-1].lower()
        if file_ext in ['jpg', 'jpeg', 'png', 'gif']:
            st.image(f"data:{attachment['type']};base64,{attachment['data']}", width=200)
        elif file_ext in ['mp4', 'avi', 'mov', 'mkv']:
            st.video(f"data:{attachment['type']};base64,{attachment['data']}")
        elif file_ext in ['mp3', 'wav', 'ogg']:
            st.audio(f"data:{attachment['type']};base64,{attachment['data']}")
        elif file_ext in ['pdf']:
            st.markdown(f"📄 [{attachment['name']}](data:{attachment['type']};base64,{attachment['data']})")
        elif file_ext in ['doc', 'docx']:
            st.markdown(f"📝 [{attachment['name']}](data:{attachment['type']};base64,{attachment['data']})")
        elif file_ext in ['xls', 'xlsx']:
            st.markdown(f"📊 [{attachment['name']}](data:{attachment['type']};base64,{attachment['data']})")
        elif file_ext in ['ppt', 'pptx']:
            st.markdown(f"📽️ [{attachment['name']}](data:{attachment['type']};base64,{attachment['data']})")
        elif file_ext in ['zip', 'rar', '7z']:
            st.markdown(f"🗜️ [{attachment['name']}](data:{attachment['type']};base64,{attachment['data']})")
        else:
            st.markdown(f"📎 [{attachment['name']}](data:{attachment['type']};base64,{attachment['data']})")

def get_file_icon(filename):
    ext = filename.split('.')[-1].lower() if '.' in filename else ''
    icons = {
        'pdf': '📄', 'doc': '📝', 'docx': '📝', 'xls': '📊', 'xlsx': '📊',
        'ppt': '📽️', 'pptx': '📽️', 'jpg': '🖼️', 'jpeg': '🖼️', 'png': '🖼️',
        'gif': '🎨', 'mp4': '🎬', 'avi': '🎬', 'mov': '🎬', 'mp3': '🎵',
        'wav': '🎵', 'zip': '🗜️', 'rar': '🗜️', 'txt': '📃', 'csv': '📋'
    }
    return icons.get(ext, '📎')

# ============ SCHOOL MANAGEMENT FUNCTIONS ============
def calculate_student_performance(grades, student_email):
    student_grades = [g for g in grades if g['student_email'] == student_email]
    if not student_grades:
        return {"average": 0, "subjects": {}, "rank": "N/A", "total_points": 0}
    
    subjects = {}
    total = 0
    for grade in student_grades:
        subjects[grade['subject']] = grade['score']
        total += grade['score']
    
    avg = total / len(student_grades)
    total_points = sum(grade['score'] for grade in student_grades)
    
    if avg >= 80:
        rank = "Excellent"
    elif avg >= 70:
        rank = "Good"
    elif avg >= 60:
        rank = "Above Average"
    elif avg >= 50:
        rank = "Average"
    elif avg >= 40:
        rank = "Below Average"
    elif avg >= 30:
        rank = "Fair"
    elif avg >= 20:
        rank = "Poor"
    else:
        rank = "Very Poor"
    
    return {"average": round(avg, 2), "subjects": subjects, "rank": rank, "total_points": total_points}

def calculate_class_performance(grades, class_name, term, year):
    class_grades = [g for g in grades if g['class_name'] == class_name and g['term'] == term and g['year'] == year]
    if not class_grades:
        return {}
    
    subject_averages = {}
    subject_counts = {}
    student_performance = {}
    
    for grade in class_grades:
        if grade['subject'] not in subject_averages:
            subject_averages[grade['subject']] = []
        subject_averages[grade['subject']].append(grade['score'])
        
        if grade['student_email'] not in student_performance:
            student_performance[grade['student_email']] = []
        student_performance[grade['student_email']].append(grade['score'])
    
    results = {}
    for subject in subject_averages:
        results[subject] = {
            "average": round(sum(subject_averages[subject]) / len(subject_averages[subject]), 2),
            "count": len(subject_averages[subject]),
            "max": max(subject_averages[subject]),
            "min": min(subject_averages[subject]),
            "median": sorted(subject_averages[subject])[len(subject_averages[subject]) // 2]
        }
    
    student_averages = [sum(scores)/len(scores) for scores in student_performance.values()]
    results["class_summary"] = {
        "total_students": len(student_performance),
        "class_average": round(sum(student_averages) / len(student_averages), 2) if student_averages else 0,
        "top_score": max(student_averages) if student_averages else 0,
        "bottom_score": min(student_averages) if student_averages else 0,
        "median_score": sorted(student_averages)[len(student_averages)//2] if student_averages else 0
    }
    
    return results

def add_academic_record(school_code, student_email, subject, score, term, year, teacher_email, class_name, exam_type="End Term"):
    grades = load_school_data(school_code, "academic_records.json", [])
    grades.append({
        "id": generate_id("GRD"),
        "student_email": student_email,
        "subject": subject,
        "score": score,
        "term": term,
        "year": year,
        "exam_type": exam_type,
        "teacher_email": teacher_email,
        "class_name": class_name,
        "date": datetime.now().strftime("%Y-%m-%d"),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })
    save_school_data(school_code, "academic_records.json", grades)
    
    # Update student performance analytics
    update_student_performance_analytics(school_code, student_email)

def update_student_performance_analytics(school_code, student_email):
    grades = load_school_data(school_code, "academic_records.json", [])
    analytics = load_school_data(school_code, "performance_analytics.json", {})
    
    student_grades = [g for g in grades if g['student_email'] == student_email]
    if student_grades:
        subjects = {}
        for grade in student_grades:
            if grade['subject'] not in subjects:
                subjects[grade['subject']] = []
            subjects[grade['subject']].append(grade['score'])
        
        subject_trends = {}
        for subject, scores in subjects.items():
            if len(scores) > 1:
                trend = scores[-1] - scores[0]
                subject_trends[subject] = {
                    "trend": "improving" if trend > 0 else "declining" if trend < 0 else "stable",
                    "change": trend,
                    "average": sum(scores)/len(scores),
                    "last_score": scores[-1]
                }
            else:
                subject_trends[subject] = {
                    "trend": "new",
                    "change": 0,
                    "average": scores[0],
                    "last_score": scores[0]
                }
        
        analytics[student_email] = {
            "subject_trends": subject_trends,
            "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_subjects": len(subjects),
            "total_exams": len(student_grades)
        }
        
        save_school_data(school_code, "performance_analytics.json", analytics)

def predict_student_performance(school_code, student_email):
    grades = load_school_data(school_code, "academic_records.json", [])
    student_grades = [g for g in grades if g['student_email'] == student_email]
    
    if len(student_grades) < 3:
        return {"error": "Insufficient data for prediction"}
    
    # Simple linear regression for prediction
    subjects = {}
    predictions = {}
    
    for grade in student_grades:
        if grade['subject'] not in subjects:
            subjects[grade['subject']] = []
        subjects[grade['subject']].append(grade['score'])
    
    for subject, scores in subjects.items():
        if len(scores) >= 3:
            # Simple trend-based prediction
            recent_trend = (scores[-1] - scores[-2]) if len(scores) >= 2 else 0
            predicted_score = scores[-1] + recent_trend
            predicted_score = max(0, min(100, predicted_score))  # Clamp between 0 and 100
            predictions[subject] = round(predicted_score, 2)
        else:
            predictions[subject] = scores[-1]
    
    return predictions

def add_attendance_record(school_code, student_email, date, status, remarks="", class_name=""):
    attendance = load_school_data(school_code, "attendance.json", [])
    attendance.append({
        "id": generate_id("ATT"),
        "student_email": student_email,
        "date": date,
        "status": status,
        "remarks": remarks,
        "class_name": class_name,
        "recorded_by": st.session_state.user['email'],
        "recorded_at": datetime.now().strftime("%Y-%m-%d %H:%M")
    })
    save_school_data(school_code, "attendance.json", attendance)

def calculate_attendance_rate(school_code, student_email, term=None, year=None):
    attendance = load_school_data(school_code, "attendance.json", [])
    student_attendance = [a for a in attendance if a['student_email'] == student_email]
    
    if term and year:
        student_attendance = [a for a in student_attendance if a['date'].startswith(f"{year}-") and a.get('term') == term]
    
    if not student_attendance:
        return 0
    
    present = len([a for a in student_attendance if a['status'] == 'Present'])
    return round((present / len(student_attendance)) * 100, 2)

def add_fee_record(school_code, student_email, amount, date, type_, status, receipt_no=None, payment_method="Cash"):
    fees = load_school_data(school_code, "fees.json", [])
    receipt = receipt_no or generate_id("RCP")
    fees.append({
        "id": generate_id("FEE"),
        "student_email": student_email,
        "amount": amount,
        "date": date,
        "type": type_,
        "status": status,
        "receipt_no": receipt,
        "payment_method": payment_method,
        "recorded_by": st.session_state.user['email'],
        "recorded_at": datetime.now().strftime("%Y-%m-%d %H:%M")
    })
    save_school_data(school_code, "fees.json", fees)
    return receipt

def calculate_fee_balance(school_code, student_email):
    fees = load_school_data(school_code, "fees.json", [])
    student_fees = [f for f in fees if f['student_email'] == student_email]
    
    total_charged = sum(f['amount'] for f in student_fees if f['type'] != 'Payment')
    total_paid = sum(f['amount'] for f in student_fees if f['status'] == 'Paid' or f['type'] == 'Payment')
    
    return total_charged - total_paid

def add_disciplinary_record(school_code, student_email, incident, action, date, recorded_by, severity="Medium"):
    discipline = load_school_data(school_code, "discipline.json", [])
    discipline.append({
        "id": generate_id("DSC"),
        "student_email": student_email,
        "incident": incident,
        "action_taken": action,
        "date": date,
        "recorded_by": recorded_by,
        "severity": severity,
        "recorded_at": datetime.now().strftime("%Y-%m-%d %H:%M")
    })
    save_school_data(school_code, "discipline.json", discipline)

def get_student_discipline_history(school_code, student_email):
    discipline = load_school_data(school_code, "discipline.json", [])
    return [d for d in discipline if d['student_email'] == student_email]

def add_teacher_review(school_code, teacher_email, student_email, review_text, rating, date, class_name):
    reviews = load_school_data(school_code, "teacher_reviews.json", [])
    reviews.append({
        "id": generate_id("REV"),
        "teacher_email": teacher_email,
        "student_email": student_email,
        "review_text": review_text,
        "rating": rating,
        "date": date,
        "class_name": class_name,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M")
    })
    save_school_data(school_code, "teacher_reviews.json", reviews)

def get_teacher_reviews(school_code, teacher_email):
    reviews = load_school_data(school_code, "teacher_reviews.json", [])
    return [r for r in reviews if r['teacher_email'] == teacher_email]

def add_parent_feedback(school_code, guardian_email, student_email, feedback_text, date):
    feedback = load_school_data(school_code, "parent_feedback.json", [])
    feedback.append({
        "id": generate_id("FDB"),
        "guardian_email": guardian_email,
        "student_email": student_email,
        "feedback_text": feedback_text,
        "date": date,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M")
    })
    save_school_data(school_code, "parent_feedback.json", feedback)

def get_parent_feedback(school_code, student_email):
    feedback = load_school_data(school_code, "parent_feedback.json", [])
    return [f for f in feedback if f['student_email'] == student_email]

# ============ TEACHER ALLOCATION FUNCTIONS ============
def allocate_teacher(school_code, teacher_name, subject, assigned_class, allocated_by):
    allocations = load_school_data(school_code, "teacher_allocations.json", [])
    allocation = {
        "id": generate_id("TAL"),
        "teacher_name": teacher_name,
        "subject": subject,
        "assigned_class": assigned_class,
        "allocated_by": allocated_by,
        "allocated_date": datetime.now().strftime("%Y-%m-%d"),
        "status": "active",
        "lessons_per_week": 0,
        "notes": ""
    }
    allocations.append(allocation)
    save_school_data(school_code, "teacher_allocations.json", allocations)
    return allocation["id"]

def get_teacher_allocations(school_code, teacher_name=None):
    allocations = load_school_data(school_code, "teacher_allocations.json", [])
    if teacher_name:
        return [a for a in allocations if a['teacher_name'] == teacher_name and a['status'] == 'active']
    return [a for a in allocations if a['status'] == 'active']

def update_teacher_allocation(school_code, allocation_id, updates):
    allocations = load_school_data(school_code, "teacher_allocations.json", [])
    for alloc in allocations:
        if alloc['id'] == allocation_id:
            alloc.update(updates)
            break
    save_school_data(school_code, "teacher_allocations.json", allocations)

def deallocate_teacher(school_code, allocation_id):
    allocations = load_school_data(school_code, "teacher_allocations.json", [])
    for alloc in allocations:
        if alloc['id'] == allocation_id:
            alloc['status'] = 'inactive'
            alloc['deallocated_date'] = datetime.now().strftime("%Y-%m-%d")
            break
    save_school_data(school_code, "teacher_allocations.json", allocations)

# ============ FURNITURE ALLOCATION FUNCTIONS ============
def allocate_furniture(school_code, item_name, location, quantity, condition, allocated_by, item_code=None):
    furniture = load_school_data(school_code, "furniture_allocations.json", [])
    code = item_code or generate_furniture_code()
    allocation = {
        "id": generate_id("FAL"),
        "item_code": code,
        "item_name": item_name,
        "location": location,
        "quantity": quantity,
        "condition": condition,
        "allocated_by": allocated_by,
        "allocated_date": datetime.now().strftime("%Y-%m-%d"),
        "status": "active",
        "last_maintenance": None,
        "next_maintenance": None,
        "notes": ""
    }
    furniture.append(allocation)
    save_school_data(school_code, "furniture_allocations.json", furniture)
    return allocation["item_code"]

def get_furniture_allocations(school_code, location=None):
    furniture = load_school_data(school_code, "furniture_allocations.json", [])
    if location:
        return [f for f in furniture if f['location'] == location and f['status'] == 'active']
    return [f for f in furniture if f['status'] == 'active']

def update_furniture_condition(school_code, item_code, condition, updated_by):
    furniture = load_school_data(school_code, "furniture_allocations.json", [])
    for item in furniture:
        if item['item_code'] == item_code:
            item['condition'] = condition
            item['last_updated'] = datetime.now().strftime("%Y-%m-%d")
            item['updated_by'] = updated_by
            break
    save_school_data(school_code, "furniture_allocations.json", furniture)

def schedule_furniture_maintenance(school_code, item_code, maintenance_date):
    furniture = load_school_data(school_code, "furniture_allocations.json", [])
    for item in furniture:
        if item['item_code'] == item_code:
            item['next_maintenance'] = maintenance_date
            break
    save_school_data(school_code, "furniture_allocations.json", furniture)

def deallocate_furniture(school_code, item_code):
    furniture = load_school_data(school_code, "furniture_allocations.json", [])
    for item in furniture:
        if item['item_code'] == item_code:
            item['status'] = 'inactive'
            item['deallocated_date'] = datetime.now().strftime("%Y-%m-%d")
            break
    save_school_data(school_code, "furniture_allocations.json", furniture)

# ============ BOOK CATALOG FUNCTIONS ============
def add_book_to_catalog(school_code, title, author, isbn, publisher, year, category, quantity, location):
    books = load_school_data(school_code, "book_catalog.json", [])
    book = {
        "id": generate_id("BOK"),
        "title": title,
        "author": author,
        "isbn": isbn,
        "publisher": publisher,
        "year": year,
        "category": category,
        "quantity": quantity,
        "available": quantity,
        "location": location,
        "added_date": datetime.now().strftime("%Y-%m-%d"),
        "added_by": st.session_state.user['email'],
        "borrowed_count": 0,
        "status": "available"
    }
    books.append(book)
    save_school_data(school_code, "book_catalog.json", books)
    return book["id"]

def borrow_book(school_code, book_id, borrower_email, borrower_name, due_date):
    books = load_school_data(school_code, "book_catalog.json", [])
    for book in books:
        if book['id'] == book_id and book['available'] > 0:
            book['available'] -= 1
            book['borrowed_count'] += 1
            if 'borrowed_by' not in book:
                book['borrowed_by'] = []
            book['borrowed_by'].append({
                "borrower_email": borrower_email,
                "borrower_name": borrower_name,
                "borrow_date": datetime.now().strftime("%Y-%m-%d"),
                "due_date": due_date,
                "returned": False
            })
            save_school_data(school_code, "book_catalog.json", books)
            return True
    return False

def return_book(school_code, book_id, borrower_email):
    books = load_school_data(school_code, "book_catalog.json", [])
    for book in books:
        if book['id'] == book_id:
            for borrow in book.get('borrowed_by', []):
                if borrow['borrower_email'] == borrower_email and not borrow['returned']:
                    borrow['returned'] = True
                    borrow['return_date'] = datetime.now().strftime("%Y-%m-%d")
                    book['available'] += 1
                    save_school_data(school_code, "book_catalog.json", books)
                    return True
    return False

def search_books(school_code, query):
    books = load_school_data(school_code, "book_catalog.json", [])
    results = []
    query = query.lower()
    for book in books:
        if (query in book['title'].lower() or 
            query in book['author'].lower() or 
            query in book.get('isbn', '').lower() or
            query in book.get('category', '').lower()):
            results.append(book)
    return results

# ============ EXAM RESULTS FUNCTIONS ============
def add_exam_result(school_code, student_email, subject, score, exam_type, term, year, class_name, max_score=100):
    results = load_school_data(school_code, "exam_results.json", [])
    percentage = (score / max_score) * 100 if max_score > 0 else 0
    grade = calculate_grade(percentage)
    
    result = {
        "id": generate_id("EXM"),
        "student_email": student_email,
        "subject": subject,
        "score": score,
        "max_score": max_score,
        "percentage": round(percentage, 2),
        "grade": grade,
        "exam_type": exam_type,
        "term": term,
        "year": year,
        "class_name": class_name,
        "recorded_by": st.session_state.user['email'],
        "recorded_date": datetime.now().strftime("%Y-%m-%d")
    }
    results.append(result)
    save_school_data(school_code, "exam_results.json", results)
    return result

def calculate_grade(percentage):
    if percentage >= 80:
        return "A"
    elif percentage >= 75:
        return "A-"
    elif percentage >= 70:
        return "B+"
    elif percentage >= 65:
        return "B"
    elif percentage >= 60:
        return "B-"
    elif percentage >= 55:
        return "C+"
    elif percentage >= 50:
        return "C"
    elif percentage >= 45:
        return "C-"
    elif percentage >= 40:
        return "D+"
    elif percentage >= 35:
        return "D"
    elif percentage >= 30:
        return "D-"
    else:
        return "E"

def get_student_exam_results(school_code, student_email, term=None, year=None):
    results = load_school_data(school_code, "exam_results.json", [])
    filtered = [r for r in results if r['student_email'] == student_email]
    if term:
        filtered = [r for r in filtered if r['term'] == term]
    if year:
        filtered = [r for r in filtered if r['year'] == year]
    return filtered

def generate_report_card(school_code, student_email, term, year):
    results = get_student_exam_results(school_code, student_email, term, year)
    if not results:
        return None
    
    subjects = {}
    total_score = 0
    total_points = 0
    
    for result in results:
        subject = result['subject']
        if subject not in subjects:
            subjects[subject] = []
        subjects[subject].append(result)
        total_score += result['percentage']
        total_points += result['score']
    
    subject_averages = {}
    for subject, exams in subjects.items():
        avg = sum(e['percentage'] for e in exams) / len(exams)
        subject_averages[subject] = {
            "average": round(avg, 2),
            "grade": calculate_grade(avg),
            "exams": len(exams),
            "best_score": max(e['score'] for e in exams),
            "worst_score": min(e['score'] for e in exams)
        }
    
    overall_avg = total_score / len(results) if results else 0
    
    report = {
        "student_email": student_email,
        "term": term,
        "year": year,
        "generated_date": datetime.now().strftime("%Y-%m-%d"),
        "subjects": subject_averages,
        "overall_average": round(overall_avg, 2),
        "overall_grade": calculate_grade(overall_avg),
        "total_exams": len(results),
        "total_score": total_score,
        "total_points": total_points
    }
    
    return report

# ============ TIMETABLE FUNCTIONS ============
def add_timetable_entry(school_code, class_name, day, time_slot, subject, teacher, room):
    timetable = load_school_data(school_code, "timetable.json", [])
    entry = {
        "id": generate_id("TT"),
        "class_name": class_name,
        "day": day,
        "time_slot": time_slot,
        "subject": subject,
        "teacher": teacher,
        "room": room,
        "created_by": st.session_state.user['email'],
        "created_date": datetime.now().strftime("%Y-%m-%d")
    }
    timetable.append(entry)
    save_school_data(school_code, "timetable.json", timetable)
    return entry

def get_class_timetable(school_code, class_name):
    timetable = load_school_data(school_code, "timetable.json", [])
    return [t for t in timetable if t['class_name'] == class_name]

def get_teacher_timetable(school_code, teacher_email):
    timetable = load_school_data(school_code, "timetable.json", [])
    return [t for t in timetable if t['teacher'] == teacher_email]

def generate_timetable_html(school_code, class_name):
    entries = get_class_timetable(school_code, class_name)
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    time_slots = sorted(set(e['time_slot'] for e in entries))
    
    html = "<table border='1'><tr><th>Time</th>"
    for day in days:
        html += f"<th>{day}</th>"
    html += "</tr>"
    
    for slot in time_slots:
        html += f"<tr><td>{slot}</td>"
        for day in days:
            entry = next((e for e in entries if e['day'] == day and e['time_slot'] == slot), None)
            if entry:
                html += f"<td>{entry['subject']}<br><small>{entry['teacher']}<br>{entry['room']}</small></td>"
            else:
                html += "<td>-</td>"
        html += "</tr>"
    
    html += "</table>"
    return html

# ============ LIBRARY FUNCTIONS ============
def add_library_book(school_code, title, author, isbn, category, shelf, quantity):
    books = load_school_data(school_code, "library_books.json", [])
    book = {
        "id": generate_id("LIB"),
        "title": title,
        "author": author,
        "isbn": isbn,
        "category": category,
        "shelf": shelf,
        "quantity": quantity,
        "available": quantity,
        "added_date": datetime.now().strftime("%Y-%m-%d"),
        "borrowed_count": 0,
        "status": "available"
    }
    books.append(book)
    save_school_data(school_code, "library_books.json", books)
    return book

def search_library(school_code, query):
    books = load_school_data(school_code, "library_books.json", [])
    results = []
    query = query.lower()
    for book in books:
        if (query in book['title'].lower() or 
            query in book['author'].lower() or 
            query in book.get('isbn', '').lower() or
            query in book.get('category', '').lower()):
            results.append(book)
    return results

def issue_library_book(school_code, book_id, user_email, user_name, due_date):
    books = load_school_data(school_code, "library_books.json", [])
    for book in books:
        if book['id'] == book_id and book['available'] > 0:
            book['available'] -= 1
            book['borrowed_count'] += 1
            if 'issued_to' not in book:
                book['issued_to'] = []
            book['issued_to'].append({
                "user_email": user_email,
                "user_name": user_name,
                "issue_date": datetime.now().strftime("%Y-%m-%d"),
                "due_date": due_date,
                "returned": False
            })
            save_school_data(school_code, "library_books.json", books)
            return True
    return False

def return_library_book(school_code, book_id, user_email):
    books = load_school_data(school_code, "library_books.json", [])
    for book in books:
        if book['id'] == book_id:
            for issue in book.get('issued_to', []):
                if issue['user_email'] == user_email and not issue['returned']:
                    issue['returned'] = True
                    issue['return_date'] = datetime.now().strftime("%Y-%m-%d")
                    book['available'] += 1
                    save_school_data(school_code, "library_books.json", books)
                    return True
    return False

# ============ SPORTS FUNCTIONS ============
def add_sport(school_code, sport_name, coach, venue, schedule):
    sports = load_school_data(school_code, "sports_data.json", [])
    sport = {
        "id": generate_id("SPT"),
        "name": sport_name,
        "coach": coach,
        "venue": venue,
        "schedule": schedule,
        "members": [],
        "created_date": datetime.now().strftime("%Y-%m-%d"),
        "created_by": st.session_state.user['email'],
        "status": "active"
    }
    sports.append(sport)
    save_school_data(school_code, "sports_data.json", sports)
    return sport

def add_sport_member(school_code, sport_id, student_email, student_name, position=None):
    sports = load_school_data(school_code, "sports_data.json", [])
    for sport in sports:
        if sport['id'] == sport_id:
            if 'members' not in sport:
                sport['members'] = []
            sport['members'].append({
                "student_email": student_email,
                "student_name": student_name,
                "position": position,
                "joined_date": datetime.now().strftime("%Y-%m-%d"),
                "status": "active"
            })
            save_school_data(school_code, "sports_data.json", sports)
            return True
    return False

def get_student_sports(school_code, student_email):
    sports = load_school_data(school_code, "sports_data.json", [])
    student_sports = []
    for sport in sports:
        for member in sport.get('members', []):
            if member['student_email'] == student_email:
                student_sports.append({
                    "sport": sport['name'],
                    "coach": sport['coach'],
                    "position": member.get('position'),
                    "joined": member['joined_date']
                })
    return student_sports

# ============ CLUBS FUNCTIONS ============
def add_club(school_code, club_name, patron, meeting_venue, meeting_schedule):
    clubs = load_school_data(school_code, "clubs_data.json", [])
    club = {
        "id": generate_id("CLB"),
        "name": club_name,
        "patron": patron,
        "venue": meeting_venue,
        "schedule": meeting_schedule,
        "members": [],
        "created_date": datetime.now().strftime("%Y-%m-%d"),
        "created_by": st.session_state.user['email'],
        "status": "active"
    }
    clubs.append(club)
    save_school_data(school_code, "clubs_data.json", clubs)
    return club

def add_club_member(school_code, club_id, student_email, student_name, role="Member"):
    clubs = load_school_data(school_code, "clubs_data.json", [])
    for club in clubs:
        if club['id'] == club_id:
            if 'members' not in club:
                club['members'] = []
            club['members'].append({
                "student_email": student_email,
                "student_name": student_name,
                "role": role,
                "joined_date": datetime.now().strftime("%Y-%m-%d"),
                "status": "active"
            })
            save_school_data(school_code, "clubs_data.json", clubs)
            return True
    return False

def get_student_clubs(school_code, student_email):
    clubs = load_school_data(school_code, "clubs_data.json", [])
    student_clubs = []
    for club in clubs:
        for member in club.get('members', []):
            if member['student_email'] == student_email:
                student_clubs.append({
                    "club": club['name'],
                    "patron": club['patron'],
                    "role": member.get('role'),
                    "joined": member['joined_date']
                })
    return student_clubs

# ============ TRANSPORT FUNCTIONS ============
def add_route(school_code, route_name, driver, vehicle, departure_time, capacity):
    transport = load_school_data(school_code, "transport_data.json", [])
    route = {
        "id": generate_id("TRN"),
        "name": route_name,
        "driver": driver,
        "vehicle": vehicle,
        "departure_time": departure_time,
        "capacity": capacity,
        "passengers": [],
        "stops": [],
        "created_date": datetime.now().strftime("%Y-%m-%d"),
        "status": "active"
    }
    transport.append(route)
    save_school_data(school_code, "transport_data.json", transport)
    return route

def add_student_to_route(school_code, route_id, student_email, student_name, pickup_point):
    transport = load_school_data(school_code, "transport_data.json", [])
    for route in transport:
        if route['id'] == route_id:
            if len(route.get('passengers', [])) < route['capacity']:
                if 'passengers' not in route:
                    route['passengers'] = []
                route['passengers'].append({
                    "student_email": student_email,
                    "student_name": student_name,
                    "pickup_point": pickup_point,
                    "added_date": datetime.now().strftime("%Y-%m-%d"),
                    "status": "active"
                })
                save_school_data(school_code, "transport_data.json", transport)
                return True
    return False

def get_student_transport(school_code, student_email):
    transport = load_school_data(school_code, "transport_data.json", [])
    for route in transport:
        for passenger in route.get('passengers', []):
            if passenger['student_email'] == student_email:
                return {
                    "route": route['name'],
                    "driver": route['driver'],
                    "vehicle": route['vehicle'],
                    "departure": route['departure_time'],
                    "pickup": passenger['pickup_point']
                }
    return None

# ============ CAFETERIA FUNCTIONS ============
def add_menu_item(school_code, item_name, category, price, available_quantity, day):
    cafeteria = load_school_data(school_code, "cafeteria_data.json", [])
    item = {
        "id": generate_id("CAF"),
        "name": item_name,
        "category": category,
        "price": price,
        "available": available_quantity,
        "day": day,
        "sold": 0,
        "added_date": datetime.now().strftime("%Y-%m-%d"),
        "status": "available"
    }
    cafeteria.append(item)
    save_school_data(school_code, "cafeteria_data.json", cafeteria)
    return item

def record_cafeteria_purchase(school_code, student_email, student_name, items, total_amount, payment_method="Cash"):
    purchases = load_school_data(school_code, "cafeteria_purchases.json", [])
    purchase = {
        "id": generate_id("PCH"),
        "student_email": student_email,
        "student_name": student_name,
        "items": items,
        "total_amount": total_amount,
        "payment_method": payment_method,
        "purchase_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "recorded_by": st.session_state.user['email']
    }
    purchases.append(purchase)
    save_school_data(school_code, "cafeteria_purchases.json", purchases)
    
    # Update cafeteria stock
    cafeteria = load_school_data(school_code, "cafeteria_data.json", [])
    for purchase_item in items:
        for item in cafeteria:
            if item['id'] == purchase_item['item_id']:
                item['sold'] += purchase_item['quantity']
                item['available'] -= purchase_item['quantity']
                break
    save_school_data(school_code, "cafeteria_data.json", cafeteria)
    
    return purchase

# ============ HEALTH RECORDS FUNCTIONS ============
def add_health_record(school_code, student_email, record_type, description, date, recorded_by):
    health = load_school_data(school_code, "health_records.json", [])
    record = {
        "id": generate_id("HLT"),
        "student_email": student_email,
        "type": record_type,
        "description": description,
        "date": date,
        "recorded_by": recorded_by,
        "recorded_at": datetime.now().strftime("%Y-%m-%d %H:%M")
    }
    health.append(record)
    save_school_data(school_code, "health_records.json", health)
    return record

def get_student_health_records(school_code, student_email):
    health = load_school_data(school_code, "health_records.json", [])
    return [h for h in health if h['student_email'] == student_email]

# ============ COUNSELING RECORDS FUNCTIONS ============
def add_counseling_session(school_code, student_email, counselor, session_type, notes, date):
    counseling = load_school_data(school_code, "counseling_records.json", [])
    session = {
        "id": generate_id("CNS"),
        "student_email": student_email,
        "counselor": counselor,
        "type": session_type,
        "notes": notes,
        "date": date,
        "recorded_at": datetime.now().strftime("%Y-%m-%d %H:%M")
    }
    counseling.append(session)
    save_school_data(school_code, "counseling_records.json", counseling)
    return session

def get_student_counseling(school_code, student_email):
    counseling = load_school_data(school_code, "counseling_records.json", [])
    return [c for c in counseling if c['student_email'] == student_email]

# ============ ALUMNI FUNCTIONS ============
def add_alumni(school_code, name, graduation_year, current_occupation, contact, achievements=""):
    alumni = load_school_data(school_code, "alumni_data.json", [])
    record = {
        "id": generate_id("ALM"),
        "name": name,
        "graduation_year": graduation_year,
        "occupation": current_occupation,
        "contact": contact,
        "achievements": achievements,
        "added_date": datetime.now().strftime("%Y-%m-%d"),
        "added_by": st.session_state.user['email'],
        "status": "active"
    }
    alumni.append(record)
    save_school_data(school_code, "alumni_data.json", alumni)
    return record

def get_alumni_by_year(school_code, year):
    alumni = load_school_data(school_code, "alumni_data.json", [])
    return [a for a in alumni if a['graduation_year'] == year]

# ============ STAFF FUNCTIONS ============
def add_staff(school_code, name, position, department, employment_date, contact, qualifications):
    staff = load_school_data(school_code, "staff_data.json", [])
    member = {
        "id": generate_id("STF"),
        "name": name,
        "position": position,
        "department": department,
        "employment_date": employment_date,
        "contact": contact,
        "qualifications": qualifications,
        "added_date": datetime.now().strftime("%Y-%m-%d"),
        "added_by": st.session_state.user['email'],
        "status": "active"
    }
    staff.append(member)
    save_school_data(school_code, "staff_data.json", staff)
    return member

def get_staff_by_department(school_code, department):
    staff = load_school_data(school_code, "staff_data.json", [])
    return [s for s in staff if s['department'] == department and s['status'] == 'active']

# ============ PARENT-TEACHER MEETINGS FUNCTIONS ============
def schedule_meeting(school_code, teacher_email, parent_email, student_email, date, time, venue, purpose):
    meetings = load_school_data(school_code, "parent_teacher_meetings.json", [])
    meeting = {
        "id": generate_id("PTM"),
        "teacher_email": teacher_email,
        "parent_email": parent_email,
        "student_email": student_email,
        "date": date,
        "time": time,
        "venue": venue,
        "purpose": purpose,
        "status": "scheduled",
        "created_by": st.session_state.user['email'],
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M")
    }
    meetings.append(meeting)
    save_school_data(school_code, "parent_teacher_meetings.json", meetings)
    return meeting

def update_meeting_status(school_code, meeting_id, status, notes=""):
    meetings = load_school_data(school_code, "parent_teacher_meetings.json", [])
    for meeting in meetings:
        if meeting['id'] == meeting_id:
            meeting['status'] = status
            meeting['notes'] = notes
            meeting['updated_at'] = datetime.now().strftime("%Y-%m-%d %H:%M")
            break
    save_school_data(school_code, "parent_teacher_meetings.json", meetings)

# ============ EXTRACURRICULAR FUNCTIONS ============
def add_activity(school_code, activity_name, category, coordinator, schedule, venue, max_participants):
    activities = load_school_data(school_code, "extracurricular_data.json", [])
    activity = {
        "id": generate_id("EXT"),
        "name": activity_name,
        "category": category,
        "coordinator": coordinator,
        "schedule": schedule,
        "venue": venue,
        "max_participants": max_participants,
        "participants": [],
        "created_date": datetime.now().strftime("%Y-%m-%d"),
        "created_by": st.session_state.user['email'],
        "status": "active"
    }
    activities.append(activity)
    save_school_data(school_code, "extracurricular_data.json", activities)
    return activity

def register_for_activity(school_code, activity_id, student_email, student_name):
    activities = load_school_data(school_code, "extracurricular_data.json", [])
    for activity in activities:
        if activity['id'] == activity_id:
            if len(activity.get('participants', [])) < activity['max_participants']:
                if 'participants' not in activity:
                    activity['participants'] = []
                activity['participants'].append({
                    "student_email": student_email,
                    "student_name": student_name,
                    "registered_date": datetime.now().strftime("%Y-%m-%d"),
                    "attendance": []
                })
                save_school_data(school_code, "extracurricular_data.json", activities)
                return True
    return False

# ============ SCHOOL CALENDAR FUNCTIONS ============
def add_calendar_event(school_code, title, event_date, event_type, description, audience, location):
    calendar = load_school_data(school_code, "school_calendar.json", [])
    event = {
        "id": generate_id("CAL"),
        "title": title,
        "date": event_date,
        "type": event_type,
        "description": description,
        "audience": audience,
        "location": location,
        "created_by": st.session_state.user['email'],
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "status": "upcoming"
    }
    calendar.append(event)
    save_school_data(school_code, "school_calendar.json", calendar)
    return event

def get_upcoming_events(school_code, days=30):
    calendar = load_school_data(school_code, "school_calendar.json", [])
    today = datetime.now().date()
    cutoff = today + timedelta(days=days)
    
    upcoming = []
    for event in calendar:
        event_date = datetime.strptime(event['date'], "%Y-%m-%d").date()
        if today <= event_date <= cutoff and event['status'] == 'upcoming':
            upcoming.append(event)
    
    return sorted(upcoming, key=lambda x: x['date'])

# ============ BUDGET FUNCTIONS ============
def add_budget_item(school_code, category, description, allocated_amount, fiscal_year):
    budget = load_school_data(school_code, "budget.json", [])
    item = {
        "id": generate_id("BDG"),
        "category": category,
        "description": description,
        "allocated": allocated_amount,
        "spent": 0,
        "remaining": allocated_amount,
        "fiscal_year": fiscal_year,
        "created_by": st.session_state.user['email'],
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "status": "active"
    }
    budget.append(item)
    save_school_data(school_code, "budget.json", budget)
    return item

def record_budget_expense(school_code, budget_id, amount, description, date):
    budget = load_school_data(school_code, "budget.json", [])
    for item in budget:
        if item['id'] == budget_id:
            item['spent'] += amount
            item['remaining'] -= amount
            if 'expenses' not in item:
                item['expenses'] = []
            item['expenses'].append({
                "amount": amount,
                "description": description,
                "date": date,
                "recorded_by": st.session_state.user['email']
            })
            save_school_data(school_code, "budget.json", budget)
            return True
    return False

# ============ INVENTORY FUNCTIONS ============
def add_inventory_item(school_code, item_name, category, quantity, unit, reorder_level, location):
    inventory = load_school_data(school_code, "inventory.json", [])
    item = {
        "id": generate_id("INV"),
        "name": item_name,
        "category": category,
        "quantity": quantity,
        "unit": unit,
        "reorder_level": reorder_level,
        "location": location,
        "added_by": st.session_state.user['email'],
        "added_date": datetime.now().strftime("%Y-%m-%d"),
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "status": "active"
    }
    inventory.append(item)
    save_school_data(school_code, "inventory.json", inventory)
    return item

def update_inventory_quantity(school_code, item_id, quantity_change, reason):
    inventory = load_school_data(school_code, "inventory.json", [])
    for item in inventory:
        if item['id'] == item_id:
            old_qty = item['quantity']
            item['quantity'] += quantity_change
            item['last_updated'] = datetime.now().strftime("%Y-%m-%d %H:%M")
            if 'transactions' not in item:
                item['transactions'] = []
            item['transactions'].append({
                "old_quantity": old_qty,
                "change": quantity_change,
                "new_quantity": item['quantity'],
                "reason": reason,
                "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "user": st.session_state.user['email']
            })
            save_school_data(school_code, "inventory.json", inventory)
            return True
    return False

def get_low_stock_items(school_code):
    inventory = load_school_data(school_code, "inventory.json", [])
    return [i for i in inventory if i['quantity'] <= i['reorder_level']]

# ============ EMAIL FUNCTIONS ============
def send_email(to_email, subject, body, from_email="school@system.com"):
    try:
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        
        # For demo purposes - in production use actual SMTP settings
        # server = smtplib.SMTP('smtp.gmail.com', 587)
        # server.starttls()
        # server.login(from_email, password)
        # server.send_message(msg)
        # server.quit()
        
        print(f"Email sent to {to_email}: {subject}")
        return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False

def send_bulk_emails(recipients, subject, body):
    success_count = 0
    for recipient in recipients:
        if send_email(recipient, subject, body):
            success_count += 1
    return success_count

# ============ QR CODE FUNCTIONS ============
def generate_qr_code_for_item(item_type, item_id, item_name):
    data = f"{item_type}:{item_id}:{item_name}"
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Convert to base64 for display
    buffered = io.BytesIO()
    img.save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue()).decode()
    
    return img_str

def generate_qr_labels(school_code, item_type, start_num, end_num):
    labels = []
    for i in range(start_num, end_num + 1):
        item_id = f"{item_type}-{i:04d}"
        item_name = f"{item_type.title()} #{i}"
        qr_data = generate_qr_code_for_item(item_type, item_id, item_name)
        labels.append({
            "item_id": item_id,
            "item_name": item_name,
            "qr_code": qr_data
        })
    return labels

# ============ REPORT GENERATION FUNCTIONS ============
def generate_pdf_report(title, data, headers):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt=title, ln=1, align='C')
    pdf.ln(10)
    
    # Add headers
    pdf.set_font("Arial", style='B', size=10)
    for header in headers:
        pdf.cell(40, 10, header, 1)
    pdf.ln()
    
    # Add data
    pdf.set_font("Arial", size=8)
    for row in data:
        for cell in row:
            pdf.cell(40, 10, str(cell), 1)
        pdf.ln()
    
    return pdf.output(dest='S').encode('latin1')

def generate_excel_report(data, sheet_name):
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    return output.getvalue()

def generate_html_report(title, data, headers):
    html = f"<html><head><title>{title}</title>"
    html += "<style>table { border-collapse: collapse; width: 100%; } th, td { border: 1px solid #ddd; padding: 8px; text-align: left; } th { background-color: #FFD700; }</style>"
    html += f"</head><body><h1>{title}</h1>"
    html += "<table><tr>"
    for header in headers:
        html += f"<th>{header}</th>"
    html += "</tr>"
    for row in data:
        html += "<tr>"
        for cell in row:
            html += f"<td>{cell}</td>"
        html += "</tr>"
    html += "</table></body></html>"
    return html

# ============ ANALYTICS FUNCTIONS ============
def analyze_performance_trends(school_code, class_name=None):
    grades = load_school_data(school_code, "academic_records.json", [])
    
    if class_name:
        grades = [g for g in grades if g['class_name'] == class_name]
    
    if not grades:
        return {}
    
    df = pd.DataFrame(grades)
    df['date'] = pd.to_datetime(df['date'])
    
    trends = {
        "overall_trend": [],
        "subject_trends": {},
        "best_performing": [],
        "needs_improvement": []
    }
    
    # Overall trend by date
    daily_avg = df.groupby('date')['score'].mean().reset_index()
    trends["overall_trend"] = daily_avg.to_dict('records')
    
    # Subject trends
    for subject in df['subject'].unique():
        subject_data = df[df['subject'] == subject]
        subject_avg = subject_data.groupby('date')['score'].mean().reset_index()
        trends["subject_trends"][subject] = subject_avg.to_dict('records')
    
    # Best performing students
    student_avg = df.groupby('student_email')['score'].mean().reset_index()
    student_avg = student_avg.sort_values('score', ascending=False).head(10)
    trends["best_performing"] = student_avg.to_dict('records')
    
    # Students needing improvement
    student_avg_low = student_avg.sort_values('score', ascending=True).head(10)
    trends["needs_improvement"] = student_avg_low.to_dict('records')
    
    return trends

def generate_performance_heatmap(school_code, term, year):
    grades = load_school_data(school_code, "academic_records.json", [])
    term_grades = [g for g in grades if g['term'] == term and g['year'] == year]
    
    if not term_grades:
        return None
    
    # Create pivot table
    df = pd.DataFrame(term_grades)
    pivot = df.pivot_table(values='score', index='student_email', columns='subject', aggfunc='mean')
    
    # Create heatmap
    fig = px.imshow(pivot,
                    title=f"Performance Heatmap - {term} {year}",
                    labels=dict(x="Subject", y="Student", color="Score"),
                    color_continuous_scale="RdYlGn")
    
    return fig

def predict_future_performance(school_code, student_email, months_ahead=3):
    grades = load_school_data(school_code, "academic_records.json", [])
    student_grades = [g for g in grades if g['student_email'] == student_email]
    
    if len(student_grades) < 3:
        return None
    
    df = pd.DataFrame(student_grades)
    df['date'] = pd.to_datetime(df['date'])
    df = df.sort_values('date')
    
    predictions = {}
    for subject in df['subject'].unique():
        subject_data = df[df['subject'] == subject]
        if len(subject_data) >= 3:
            # Simple moving average prediction
            recent_scores = subject_data['score'].tail(3).tolist()
            avg_trend = (recent_scores[-1] - recent_scores[0]) / len(recent_scores)
            last_score = recent_scores[-1]
            predicted = last_score + (avg_trend * months_ahead)
            predicted = max(0, min(100, predicted))
            predictions[subject] = round(predicted, 2)
    
    return predictions

# ============ SESSION STATE ============
if 'user' not in st.session_state:
    st.session_state.user = None
if 'current_school' not in st.session_state:
    st.session_state.current_school = None
if 'page' not in st.session_state:
    st.session_state.page = 'welcome'
if 'menu_index' not in st.session_state:
    st.session_state.menu_index = 0
if 'chat_with' not in st.session_state:
    st.session_state.chat_with = None
if 'group_chat_with' not in st.session_state:
    st.session_state.group_chat_with = None
if 'main_section' not in st.session_state:
    st.session_state.main_section = 'community'
if 'management_access' not in st.session_state:
    st.session_state.management_access = False
if 'personal_access' not in st.session_state:
    st.session_state.personal_access = False

# ============ MAIN APP ============

# ----- WELCOME PAGE WITH THREE MAIN BUTTONS -----
if st.session_state.page == 'welcome':
    st.markdown('<h1>✨ Complete School Community Hub ✨</h1>', unsafe_allow_html=True)
    st.markdown('<p style="text-align: center; color: #333333; font-size: 1.2rem;">Connect • Manage • Personalize • Analyze</p>', unsafe_allow_html=True)
    st.divider()
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("🏫 School Community", key="nav_community", use_container_width=True):
            st.session_state.main_section = 'community'
            st.rerun()
        st.markdown("""
        <p style="text-align: center; color: #666; padding: 10px;">Connect with teachers, students, and guardians. Join groups, chat, and collaborate!</p>
        """, unsafe_allow_html=True)
    
    with col2:
        if st.button("📊 School Management", key="nav_management", use_container_width=True):
            st.session_state.main_section = 'management'
            st.rerun()
        st.markdown("""
        <p style="text-align: center; color: #666; padding: 10px;">Complete school administration - Academics, Finance, Discipline, Teacher/Furniture Allocation, Inventory, Library, Transport, Cafeteria, Health, Counseling, Alumni, Staff, Budget, Procurement, Assets, Maintenance, Security, Emergency, Insurance, Sponsorship, Scholarships, Donations, Grants, Partnerships, MOUs, Visitors, Events, Trips, Competitions, Awards, Certificates, Transcripts, Reports, Analytics, Predictions, Recommendations, Feedback, Surveys, Polls, Voting, Elections, Nominations, Applications, Interviews, Recruitment, Training, Workshops, Seminars, Conferences, Webinars, Online Courses, E-Learning, Digital Library, Research, Publications, Journals, Articles, Blogs, Forums, Discussions, Q&A, Help Desk, Support Tickets, Tech Support, IT Assets, Software Licenses, Hardware, Network, Servers, Backups, Disaster Recovery, Business Continuity, Risk Management, Compliance, Audit, Quality Assurance, Performance Reviews, Appraisals, Promotions, Transfers, Retirements, Resignations, Terminations, Leave Requests, Holidays, Attendance Requests, Overtime, Payroll, Salaries, Allowances, Deductions, Benefits, Taxes, Pensions, Loans, Advances, Claims, Reimbursements, Petty Cash, Bank Accounts, Cash Flow, Investments, Assets & Liabilities, Equity, Revenue, Expenses, Profit & Loss, Balance Sheet, Financial Ratios, Budget Variance, Forecasts, Projections, Scenarios, Strategic Plan, Vision & Mission, Core Values, SWOT Analysis, PESTLE Analysis, Competitor Analysis, Market Research, Branding, Marketing, Social Media, Website, Newsletters, Press Releases, Media Coverage, Public Relations, Community Outreach, Corporate Social Responsibility, Sustainability, Environmental Impact, Carbon Footprint, Green Initiatives, Waste Management, Recycling, Energy Efficiency, Water Conservation, Renewable Energy</p>
        """, unsafe_allow_html=True)
    
    with col3:
        if st.button("👤 Personal Dashboard", key="nav_personal", use_container_width=True):
            st.session_state.main_section = 'personal'
            st.rerun()
        st.markdown("""
        <p style="text-align: center; color: #666; padding: 10px;">Your personal information, performance, reviews, class analytics, attendance, fee status, discipline history, health records, counseling sessions, extracurricular activities, sports, clubs, transport, cafeteria purchases, alumni connections, and personalized recommendations!</p>
        """, unsafe_allow_html=True)
    
    st.divider()
    
    # ============ SCHOOL COMMUNITY SECTION (PUBLIC) ============
    if st.session_state.main_section == 'community':
        st.markdown('<div class="section-container section-community">', unsafe_allow_html=True)
        st.markdown("""
        <div style="background: linear-gradient(135deg, #FF6B6B, #FF8E8E); padding: 30px; border-radius: 20px; text-align: center; margin-bottom: 30px; box-shadow: 0 0 40px rgba(255, 107, 107, 0.5);">
            <h3 style="color: white; text-shadow: 2px 2px 4px rgba(0,0,0,0.3);">🏫 School Community</h3>
            <p style="color: white;">Login or register to connect with your school community!</p>
        </div>
        """, unsafe_allow_html=True)
        
        tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
            "👑 Admin Login", "🏫 Create School", "👨‍🏫 Teacher", "👨‍🎓 Student", 
            "👪 Guardian", "🤝 Alumni", "👥 Visitor"
        ])
        
        with tab1:
            col1, col2 = st.columns([1,1])
            with col1:
                with st.form("admin_login"):
                    st.subheader("Admin Login")
                    school_code = st.text_input("School Code", placeholder="Enter your school code")
                    admin_email = st.text_input("Email", placeholder="admin@school.edu")
                    admin_password = st.text_input("Password", type="password", placeholder="••••••••")
                    if st.form_submit_button("Login", use_container_width=True):
                        if not school_code or not admin_email or not admin_password:
                            st.error("Please fill all fields")
                        else:
                            all_schools = load_all_schools()
                            if school_code in all_schools:
                                school = all_schools[school_code]
                                if school['admin_email'] == admin_email:
                                    users = load_school_data(school_code, "users.json", [])
                                    for u in users:
                                        if u['email'] == admin_email and verify_password(admin_password, u['password']) and u['role'] == 'admin':
                                            st.session_state.current_school = school
                                            st.session_state.user = u
                                            st.session_state.management_access = True
                                            st.session_state.personal_access = True
                                            st.session_state.page = 'dashboard'
                                            st.rerun()
                                    st.error("Invalid password")
                                else:
                                    st.error("Not the admin email")
                            else:
                                st.error("School not found")
        
        with tab2:
            col1, col2 = st.columns([1,1])
            with col1:
                with st.form("create_school"):
                    st.subheader("Create a New School")
                    school_name = st.text_input("School Name", placeholder="e.g., Golden Heights Academy")
                    admin_name = st.text_input("Your Full Name", placeholder="e.g., John Doe")
                    admin_email = st.text_input("Your Email", placeholder="you@school.edu")
                    password = st.text_input("Password", type="password", placeholder="Create a strong password")
                    confirm = st.text_input("Confirm Password", type="password", placeholder="Re-enter password")
                    city = st.text_input("City", placeholder="e.g., Nairobi")
                    state = st.text_input("State/Province", placeholder="e.g., Nairobi")
                    motto = st.text_input("School Motto", placeholder="e.g., Excellence is Our Tradition")
                    
                    if st.form_submit_button("Create School", use_container_width=True):
                        if not school_name or not admin_email or not password:
                            st.error("School name, email and password are required")
                        elif password != confirm:
                            st.error("Passwords do not match")
                        else:
                            all_schools = load_all_schools()
                            code = generate_school_code()
                            while code in all_schools:
                                code = generate_school_code()
                            
                            new_school = {
                                "code": code,
                                "name": school_name,
                                "city": city,
                                "state": state,
                                "motto": motto,
                                "created": datetime.now().strftime("%Y-%m-%d"),
                                "admin_email": admin_email,
                                "admin_name": admin_name,
                                "stats": {"students":0, "teachers":0, "guardians":0, "classes":0, "alumni":0}
                            }
                            all_schools[code] = new_school
                            save_all_schools(all_schools)
                            
                            users = [{
                                "user_id": generate_id("USR"),
                                "email": admin_email,
                                "fullname": admin_name,
                                "password": hash_password(password),
                                "role": "admin",
                                "joined": datetime.now().strftime("%Y-%m-%d"),
                                "school_code": code,
                                "class": "Administration",
                                "stream": "",
                                "admission_number": "",
                                "profile_pic": None,
                                "bio": "",
                                "phone": ""
                            }]
                            save_school_data(code, "users.json", users)
                            save_school_data(code, "classes.json", [])
                            save_school_data(code, "groups.json", [])
                            save_school_data(code, "announcements.json", [])
                            save_school_data(code, "assignments.json", [])
                            save_school_data(code, "messages.json", [])
                            save_school_data(code, "friend_requests.json", [])
                            save_school_data(code, "friendships.json", [])
                            save_school_data(code, "group_chats.json", [])
                            save_school_data(code, "academic_records.json", [])
                            save_school_data(code, "attendance.json", [])
                            save_school_data(code, "fees.json", [])
                            save_school_data(code, "discipline.json", [])
                            save_school_data(code, "teacher_reviews.json", [])
                            save_school_data(code, "parent_feedback.json", [])
                            save_school_data(code, "class_requests.json", [])
                            save_school_data(code, "group_requests.json", [])
                            save_school_data(code, "teacher_allocations.json", [])
                            save_school_data(code, "furniture_allocations.json", [])
                            save_school_data(code, "book_catalog.json", [])
                            save_school_data(code, "exam_results.json", [])
                            save_school_data(code, "timetable.json", [])
                            save_school_data(code, "library_books.json", [])
                            save_school_data(code, "sports_data.json", [])
                            save_school_data(code, "clubs_data.json", [])
                            save_school_data(code, "transport_data.json", [])
                            save_school_data(code, "cafeteria_data.json", [])
                            save_school_data(code, "cafeteria_purchases.json", [])
                            save_school_data(code, "health_records.json", [])
                            save_school_data(code, "counseling_records.json", [])
                            save_school_data(code, "alumni_data.json", [])
                            save_school_data(code, "staff_data.json", [])
                            save_school_data(code, "parent_teacher_meetings.json", [])
                            save_school_data(code, "extracurricular_data.json", [])
                            save_school_data(code, "school_calendar.json", [])
                            save_school_data(code, "budget.json", [])
                            save_school_data(code, "inventory.json", [])
                            
                            st.session_state.current_school = new_school
                            st.session_state.user = users[0]
                            st.session_state.management_access = True
                            st.session_state.personal_access = True
                            st.session_state.page = 'dashboard'
                            st.success(f"✅ School Created! Your School Code is: **{code}**")
                            st.info("Save this code - you'll need it for login!")
                            st.rerun()
        
        with tab3:
            subtab1, subtab2 = st.tabs(["Login", "Register"])
            with subtab1:
                with st.form("teacher_login"):
                    st.subheader("Teacher Login")
                    school_code = st.text_input("School Code")
                    email = st.text_input("Email")
                    password = st.text_input("Password", type="password")
                    if st.form_submit_button("Login", use_container_width=True):
                        if not school_code or not email or not password:
                            st.error("All fields required")
                        else:
                            all_schools = load_all_schools()
                            if school_code in all_schools:
                                school = all_schools[school_code]
                                users = load_school_data(school_code, "users.json", [])
                                for u in users:
                                    if u['email'] == email and verify_password(password, u['password']) and u['role'] == 'teacher':
                                        st.session_state.current_school = school
                                        st.session_state.user = u
                                        st.session_state.management_access = True
                                        st.session_state.personal_access = True
                                        st.session_state.page = 'dashboard'
                                        st.rerun()
                                st.error("Invalid credentials")
                            else:
                                st.error("School not found")
            
            with subtab2:
                with st.form("teacher_register"):
                    st.subheader("New Teacher Registration")
                    school_code = st.text_input("School Code")
                    teacher_code = st.text_input("Teacher Code (from admin)")
                    fullname = st.text_input("Full Name")
                    email = st.text_input("Email")
                    password = st.text_input("Password", type="password")
                    confirm = st.text_input("Confirm Password", type="password")
                    
                    if st.form_submit_button("Register", use_container_width=True):
                        if not all([school_code, teacher_code, fullname, email, password]):
                            st.error("All fields required")
                        elif password != confirm:
                            st.error("Passwords don't match")
                        else:
                            all_schools = load_all_schools()
                            if school_code not in all_schools:
                                st.error("School not found")
                            else:
                                school = all_schools[school_code]
                                users = load_school_data(school_code, "users.json", [])
                                
                                if any(u['email'] == email for u in users):
                                    st.error("❌ Email already registered!")
                                    st.stop()
                                
                                new_user = {
                                    "user_id": generate_id("USR"),
                                    "email": email,
                                    "fullname": fullname,
                                    "password": hash_password(password),
                                    "role": "teacher",
                                    "joined": datetime.now().strftime("%Y-%m-%d"),
                                    "school_code": school_code,
                                    "teacher_code": teacher_code,
                                    "class": "",
                                    "stream": "",
                                    "admission_number": "",
                                    "profile_pic": None,
                                    "bio": "",
                                    "phone": ""
                                }
                                users.append(new_user)
                                save_school_data(school_code, "users.json", users)
                                school['stats']['teachers'] = school['stats'].get('teachers', 0) + 1
                                all_schools[school_code] = school
                                save_all_schools(all_schools)
                                
                                st.session_state.current_school = school
                                st.session_state.user = new_user
                                st.session_state.management_access = True
                                st.session_state.personal_access = True
                                st.session_state.page = 'dashboard'
                                st.success("✅ Registration Successful!")
                                st.rerun()
        
        with tab4:
            subtab1, subtab2 = st.tabs(["Login", "Register"])
            with subtab1:
                with st.form("student_login"):
                    st.subheader("Student Login")
                    school_code = st.text_input("School Code")
                    admission_number = st.text_input("Admission Number")
                    password = st.text_input("Password", type="password")
                    if st.form_submit_button("Login", use_container_width=True):
                        if not school_code or not admission_number or not password:
                            st.error("All fields required")
                        else:
                            all_schools = load_all_schools()
                            if school_code in all_schools:
                                school = all_schools[school_code]
                                users = load_school_data(school_code, "users.json", [])
                                for u in users:
                                    if u.get('admission_number') == admission_number and verify_password(password, u['password']) and u['role'] == 'student':
                                        st.session_state.current_school = school
                                        st.session_state.user = u
                                        st.session_state.management_access = False
                                        st.session_state.personal_access = True
                                        st.session_state.page = 'dashboard'
                                        st.rerun()
                                st.error("Invalid admission number or password")
                            else:
                                st.error("School not found")
            
            with subtab2:
                with st.form("student_register"):
                    st.subheader("New Student Registration")
                    school_code = st.text_input("School Code")
                    fullname = st.text_input("Full Name")
                    email = st.text_input("Email (Optional)")
                    class_name = st.selectbox("Class", KENYAN_GRADES)
                    stream = st.selectbox("Stream", STREAMS)
                    password = st.text_input("Password", type="password")
                    confirm = st.text_input("Confirm Password", type="password")
                    
                    if st.form_submit_button("Register", use_container_width=True):
                        if not all([school_code, fullname, password]):
                            st.error("School code, name and password are required")
                        elif password != confirm:
                            st.error("Passwords don't match")
                        else:
                            all_schools = load_all_schools()
                            if school_code not in all_schools:
                                st.error("School not found")
                            else:
                                school = all_schools[school_code]
                                users = load_school_data(school_code, "users.json", [])
                                
                                if email and any(u['email'] == email for u in users):
                                    st.error("❌ Email already registered!")
                                    st.stop()
                                
                                admission_number = generate_admission_number()
                                while any(u.get('admission_number') == admission_number for u in users):
                                    admission_number = generate_admission_number()
                                
                                new_user = {
                                    "user_id": generate_id("USR"),
                                    "email": email if email else "",
                                    "fullname": fullname,
                                    "password": hash_password(password),
                                    "role": "student",
                                    "joined": datetime.now().strftime("%Y-%m-%d"),
                                    "school_code": school_code,
                                    "class": class_name,
                                    "stream": stream,
                                    "admission_number": admission_number,
                                    "guardians": [],
                                    "profile_pic": None,
                                    "bio": "",
                                    "phone": ""
                                }
                                users.append(new_user)
                                save_school_data(school_code, "users.json", users)
                                school['stats']['students'] = school['stats'].get('students', 0) + 1
                                all_schools[school_code] = school
                                save_all_schools(all_schools)
                                
                                st.success(f"✅ Registered! Your Admission Number is: **{admission_number}**")
                                st.info("📝 Save this number - you'll need it to login!")
        
        with tab5:
            subtab1, subtab2 = st.tabs(["Login", "Register"])
            with subtab1:
                with st.form("guardian_login"):
                    st.subheader("Guardian Login")
                    school_code = st.text_input("School Code")
                    student_admission = st.text_input("Student's Admission Number")
                    email = st.text_input("Your Email")
                    password = st.text_input("Password", type="password")
                    if st.form_submit_button("Login", use_container_width=True):
                        if not school_code or not student_admission or not email or not password:
                            st.error("All fields required")
                        else:
                            all_schools = load_all_schools()
                            if school_code in all_schools:
                                school = all_schools[school_code]
                                users = load_school_data(school_code, "users.json", [])
                                for u in users:
                                    if u['role'] == 'guardian' and u['email'] == email and verify_password(password, u['password']):
                                        if student_admission in u.get('linked_students', []):
                                            st.session_state.current_school = school
                                            st.session_state.user = u
                                            st.session_state.management_access = False
                                            st.session_state.personal_access = True
                                            st.session_state.page = 'dashboard'
                                            st.rerun()
                                        else:
                                            st.error("You are not linked to this student")
                                            break
                                st.error("Invalid credentials")
                            else:
                                st.error("School not found")
            
            with subtab2:
                with st.form("guardian_register"):
                    st.subheader("New Guardian Registration")
                    school_code = st.text_input("School Code")
                    student_admission = st.text_input("Student's Admission Number")
                    fullname = st.text_input("Your Full Name")
                    email = st.text_input("Your Email")
                    phone = st.text_input("Phone Number")
                    password = st.text_input("Password", type="password")
                    confirm = st.text_input("Confirm Password", type="password")
                    
                    if st.form_submit_button("Register", use_container_width=True):
                        if not all([school_code, student_admission, fullname, email, password]):
                            st.error("All fields required")
                        elif password != confirm:
                            st.error("Passwords don't match")
                        else:
                            all_schools = load_all_schools()
                            if school_code not in all_schools:
                                st.error("School not found")
                            else:
                                school = all_schools[school_code]
                                users = load_school_data(school_code, "users.json", [])
                                
                                if any(u['email'] == email for u in users):
                                    st.error("❌ Email already registered!")
                                    st.stop()
                                
                                student = None
                                for u in users:
                                    if u.get('admission_number') == student_admission and u['role'] == 'student':
                                        student = u
                                        break
                                
                                if not student:
                                    st.error("❌ Student not found with this admission number!")
                                    st.stop()
                                
                                new_user = {
                                    "user_id": generate_id("USR"),
                                    "email": email,
                                    "fullname": fullname,
                                    "phone": phone,
                                    "password": hash_password(password),
                                    "role": "guardian",
                                    "joined": datetime.now().strftime("%Y-%m-%d"),
                                    "school_code": school_code,
                                    "linked_students": [student_admission],
                                    "class": "",
                                    "stream": "",
                                    "admission_number": "",
                                    "profile_pic": None,
                                    "bio": "",
                                }
                                users.append(new_user)
                                
                                if 'guardians' not in student:
                                    student['guardians'] = []
                                student['guardians'].append(email)
                                
                                save_school_data(school_code, "users.json", users)
                                school['stats']['guardians'] = school['stats'].get('guardians', 0) + 1
                                all_schools[school_code] = school
                                save_all_schools(all_schools)
                                
                                st.success("✅ Guardian Registration Successful!")
        
        with tab6:
            st.subheader("Alumni Registration")
            with st.form("alumni_register"):
                school_code = st.text_input("School Code")
                fullname = st.text_input("Full Name")
                graduation_year = st.number_input("Graduation Year", min_value=1950, max_value=datetime.now().year, value=datetime.now().year-1)
                occupation = st.text_input("Current Occupation")
                email = st.text_input("Email")
                phone = st.text_input("Phone")
                achievements = st.text_area("Achievements (Optional)")
                
                if st.form_submit_button("Register as Alumni"):
                    if not all([school_code, fullname, graduation_year, occupation, email]):
                        st.error("All fields required")
                    else:
                        all_schools = load_all_schools()
                        if school_code not in all_schools:
                            st.error("School not found")
                        else:
                            add_alumni(school_code, fullname, str(graduation_year), occupation, email, achievements)
                            st.success("✅ Alumni registration successful!")
        
        with tab7:
            st.subheader("Visitor Access")
            st.info("As a visitor, you can view public announcements and school information.")
            if st.button("Continue as Visitor"):
                st.session_state.user = {
                    "user_id": "visitor",
                    "fullname": "Guest Visitor",
                    "role": "visitor",
                    "email": "visitor@guest.com"
                }
                st.session_state.page = 'dashboard'
                st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # ============ SCHOOL MANAGEMENT SECTION ============
    elif st.session_state.main_section == 'management':
        st.markdown('<div class="section-container section-management">', unsafe_allow_html=True)
        st.markdown("""
        <div style="background: linear-gradient(135deg, #4ECDC4, #6EE7E7); padding: 30px; border-radius: 20px; text-align: center; margin-bottom: 30px; box-shadow: 0 0 40px rgba(78, 205, 196, 0.5);">
            <h3 style="color: white; text-shadow: 2px 2px 4px rgba(0,0,0,0.3);">📊 School Management System</h3>
            <p style="color: white;">Complete school administration - Requires login with admin/teacher credentials</p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.session_state.user and st.session_state.current_school:
            if st.session_state.user['role'] in ['admin', 'teacher']:
                school_code = st.session_state.current_school['code']
                users = load_school_data(school_code, "users.json", [])
                
                mgmt_tabs = st.tabs([
                    "📚 Academics", "💰 Finance", "📋 Discipline", "👨‍🏫 Teacher Allocation", 
                    "🪑 Furniture", "📖 Library", "🏃 Sports", "🎭 Clubs", "🚌 Transport", 
                    "🍽️ Cafeteria", "🏥 Health", "🧠 Counseling", "🎓 Alumni", "👥 Staff", 
                    "📅 Calendar", "📊 Reports", "📈 Analytics", "🔮 Predictions", 
                    "📦 Inventory", "💰 Budget", "⚙️ Settings"
                ])
                
                with mgmt_tabs[0]:  # Academics
                    st.subheader("Academic Records")
                    
                    # Add Academic Record
                    with st.expander("➕ Add New Academic Record"):
                        students = [u for u in users if u['role'] == 'student']
                        with st.form("add_academic"):
                            student = st.selectbox("Select Student", [f"{s['fullname']} ({s['admission_number']}) - {s['class']}" for s in students])
                            subject = st.selectbox("Subject", PRIMARY_SUBJECTS)
                            exam_type = st.selectbox("Exam Type", EXAM_TYPES)
                            score = st.number_input("Score (0-100)", 0, 100, 0)
                            term = st.selectbox("Term", TERMS)
                            year = st.number_input("Year", datetime.now().year)
                            
                            if st.form_submit_button("Save Record"):
                                student_email = student.split('(')[1].split(')')[0]
                                student_obj = next(s for s in students if s['email'] == student_email)
                                add_academic_record(
                                    school_code, student_email, subject, score, term, str(year),
                                    st.session_state.user['email'], student_obj['class'], exam_type
                                )
                                st.success("Record added!")
                                st.rerun()
                    
                    # View Records
                    st.subheader("Recent Academic Records")
                    grades = load_school_data(school_code, "academic_records.json", [])
                    if grades:
                        df = pd.DataFrame(grades[-20:])
                        st.dataframe(df[['student_email', 'subject', 'score', 'term', 'year', 'exam_type']])
                    
                    # Performance Analytics
                    st.subheader("Performance Analytics")
                    trends = analyze_performance_trends(school_code)
                    if trends:
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("Total Records", len(grades))
                            st.metric("Average Score", round(sum(g['score'] for g in grades)/len(grades), 2) if grades else 0)
                        with col2:
                            st.metric("Best Student", trends['best_performing'][0]['student_email'] if trends['best_performing'] else "N/A")
                            st.metric("Worst Student", trends['needs_improvement'][0]['student_email'] if trends['needs_improvement'] else "N/A")
                
                with mgmt_tabs[1]:  # Finance
                    st.subheader("Fee Management")
                    
                    with st.expander("➕ Add Fee Record"):
                        students = [u for u in users if u['role'] == 'student']
                        with st.form("add_fee"):
                            student = st.selectbox("Select Student", [f"{s['fullname']} ({s['admission_number']})" for s in students])
                            amount = st.number_input("Amount (KES)", 0.0, step=100.0)
                            fee_type = st.selectbox("Type", ["Tuition", "Transport", "Lunch", "Development", "Uniform", "Activity"])
                            status = st.selectbox("Status", ["Paid", "Pending", "Overdue"])
                            payment_method = st.selectbox("Payment Method", ["Cash", "M-Pesa", "Bank Transfer", "Cheque"])
                            
                            if st.form_submit_button("Save Fee"):
                                student_email = student.split('(')[1].rstrip(')')
                                receipt = add_fee_record(
                                    school_code, student_email, amount,
                                    datetime.now().strftime("%Y-%m-%d"), fee_type, status,
                                    payment_method=payment_method
                                )
                                st.success(f"Fee recorded! Receipt No: {receipt}")
                                st.rerun()
                    
                    # Fee Summary
                    fees = load_school_data(school_code, "fees.json", [])
                    if fees:
                        total_collected = sum(f['amount'] for f in fees if f['status'] == 'Paid')
                        total_pending = sum(f['amount'] for f in fees if f['status'] in ['Pending', 'Overdue'])
                        
                        col1, col2, col3 = st.columns(3)
                        col1.metric("Total Collected", f"KES {total_collected:,.0f}")
                        col2.metric("Total Pending", f"KES {total_pending:,.0f}")
                        col3.metric("Collection Rate", f"{total_collected/(total_collected+total_pending)*100:.1f}%" if (total_collected+total_pending)>0 else "0%")
                        
                        # Fee by type chart
                        fee_by_type = {}
                        for fee in fees:
                            fee_by_type[fee['type']] = fee_by_type.get(fee['type'], 0) + fee['amount']
                        
                        df = pd.DataFrame(list(fee_by_type.items()), columns=['Type', 'Amount'])
                        fig = px.pie(df, values='Amount', names='Type', title='Fees by Type')
                        st.plotly_chart(fig)
                
                with mgmt_tabs[2]:  # Discipline
                    st.subheader("Discipline Records")
                    
                    with st.expander("➕ Add Discipline Record"):
                        students = [u for u in users if u['role'] == 'student']
                        with st.form("add_discipline"):
                            student = st.selectbox("Select Student", [f"{s['fullname']} ({s['admission_number']})" for s in students])
                            incident = st.text_area("Incident")
                            action = st.text_area("Action Taken")
                            severity = st.selectbox("Severity", ["Low", "Medium", "High", "Critical"])
                            
                            if st.form_submit_button("Save Record"):
                                student_email = student.split('(')[1].rstrip(')')
                                add_disciplinary_record(
                                    school_code, student_email, incident, action,
                                    datetime.now().strftime("%Y-%m-%d"), st.session_state.user['email'], severity
                                )
                                st.success("Record saved!")
                                st.rerun()
                    
                    # View Discipline Records
                    discipline = load_school_data(school_code, "discipline.json", [])
                    if discipline:
                        df = pd.DataFrame(discipline)
                        st.dataframe(df[['student_email', 'incident', 'severity', 'date']])
                
                with mgmt_tabs[3]:  # Teacher Allocation
                    st.subheader("Teacher Allocation")
                    
                    with st.expander("➕ Allocate Teacher"):
                        teachers = [u for u in users if u['role'] == 'teacher']
                        with st.form("allocate_teacher"):
                            teacher = st.selectbox("Select Teacher", [f"{t['fullname']} ({t['email']})" for t in teachers])
                            subject = st.selectbox("Subject", PRIMARY_SUBJECTS)
                            assigned_class = st.selectbox("Assigned Class", KENYAN_GRADES)
                            lessons = st.number_input("Lessons per Week", 1, 40, 8)
                            
                            if st.form_submit_button("Allocate"):
                                teacher_name = teacher.split('(')[0].strip()
                                allocation_id = allocate_teacher(
                                    school_code, teacher_name, subject, assigned_class,
                                    st.session_state.user['email']
                                )
                                update_teacher_allocation(school_code, allocation_id, {"lessons_per_week": lessons})
                                st.success("Teacher allocated!")
                                st.rerun()
                    
                    # View Allocations
                    allocations = get_teacher_allocations(school_code)
                    if allocations:
                        df = pd.DataFrame(allocations)
                        st.dataframe(df[['teacher_name', 'subject', 'assigned_class', 'lessons_per_week']])
                
                with mgmt_tabs[4]:  # Furniture
                    st.subheader("Furniture Allocation")
                    
                    with st.expander("➕ Add Furniture"):
                        with st.form("add_furniture"):
                            item = st.text_input("Item Name")
                            location = st.text_input("Location/Room")
                            quantity = st.number_input("Quantity", 1, 1000, 1)
                            condition = st.selectbox("Condition", ["Good", "Fair", "Broken", "Under Maintenance"])
                            
                            if st.form_submit_button("Add Furniture"):
                                code = allocate_furniture(
                                    school_code, item, location, quantity, condition,
                                    st.session_state.user['email']
                                )
                                st.success(f"Furniture added! Code: {code}")
                                st.rerun()
                    
                    # View Furniture
                    furniture = get_furniture_allocations(school_code)
                    if furniture:
                        df = pd.DataFrame(furniture)
                        st.dataframe(df[['item_code', 'item_name', 'location', 'quantity', 'condition']])
                        
                        # Low stock warning
                        low_stock = [f for f in furniture if f['condition'] == 'Broken']
                        if low_stock:
                            st.warning(f"⚠️ {len(low_stock)} items need maintenance!")
                
                with mgmt_tabs[5]:  # Library
                    st.subheader("Library Management")
                    
                    tab_a, tab_b, tab_c = st.tabs(["Add Book", "Search Books", "Issue/Return"])
                    
                    with tab_a:
                        with st.form("add_book"):
                            title = st.text_input("Title")
                            author = st.text_input("Author")
                            isbn = st.text_input("ISBN")
                            category = st.selectbox("Category", ["Textbook", "Novel", "Reference", "Journal", "Magazine"])
                            quantity = st.number_input("Quantity", 1, 100, 1)
                            location = st.text_input("Shelf Location")
                            
                            if st.form_submit_button("Add Book"):
                                add_library_book(school_code, title, author, isbn, category, location, quantity)
                                st.success("Book added!")
                                st.rerun()
                    
                    with tab_b:
                        search = st.text_input("Search books by title, author, or ISBN")
                        if search:
                            results = search_library(school_code, search)
                            if results:
                                for book in results:
                                    st.markdown(f"""
                                    <div class="golden-card">
                                        <strong>{book['title']}</strong> by {book['author']}<br>
                                        ISBN: {book['isbn']} | Category: {book['category']}<br>
                                        Available: {book['available']}/{book['quantity']} | Location: {book['shelf']}
                                    </div>
                                    """, unsafe_allow_html=True)
                            else:
                                st.info("No books found")
                    
                    with tab_c:
                        col1, col2 = st.columns(2)
                        with col1:
                            with st.form("issue_book"):
                                book_id = st.text_input("Book ID")
                                user_email = st.text_input("User Email")
                                user_name = st.text_input("User Name")
                                due_date = st.date_input("Due Date", datetime.now() + timedelta(days=14))
                                
                                if st.form_submit_button("Issue Book"):
                                    if issue_library_book(school_code, book_id, user_email, user_name, due_date.strftime("%Y-%m-%d")):
                                        st.success("Book issued!")
                                    else:
                                        st.error("Failed to issue book")
                        
                        with col2:
                            with st.form("return_book"):
                                book_id_return = st.text_input("Book ID to Return")
                                user_email_return = st.text_input("User Email for Return")
                                
                                if st.form_submit_button("Return Book"):
                                    if return_library_book(school_code, book_id_return, user_email_return):
                                        st.success("Book returned!")
                                    else:
                                        st.error("Failed to return book")
                
                with mgmt_tabs[6]:  # Sports
                    st.subheader("Sports Management")
                    
                    with st.expander("➕ Add Sport"):
                        with st.form("add_sport"):
                            sport_name = st.text_input("Sport Name")
                            coach = st.text_input("Coach Name")
                            venue = st.text_input("Venue")
                            schedule = st.text_input("Schedule (e.g., Mon/Wed 4PM)")
                            
                            if st.form_submit_button("Add Sport"):
                                add_sport(school_code, sport_name, coach, venue, schedule)
                                st.success("Sport added!")
                                st.rerun()
                    
                    sports = load_school_data(school_code, "sports_data.json", [])
                    if sports:
                        for sport in sports:
                            with st.expander(f"🏅 {sport['name']}"):
                                st.write(f"**Coach:** {sport['coach']}")
                                st.write(f"**Venue:** {sport['venue']}")
                                st.write(f"**Schedule:** {sport['schedule']}")
                                st.write(f"**Members:** {len(sport.get('members', []))}")
                
                with mgmt_tabs[7]:  # Clubs
                    st.subheader("Clubs Management")
                    
                    with st.expander("➕ Add Club"):
                        with st.form("add_club"):
                            club_name = st.text_input("Club Name")
                            patron = st.text_input("Patron Name")
                            venue = st.text_input("Meeting Venue")
                            schedule = st.text_input("Meeting Schedule")
                            
                            if st.form_submit_button("Add Club"):
                                add_club(school_code, club_name, patron, venue, schedule)
                                st.success("Club added!")
                                st.rerun()
                    
                    clubs = load_school_data(school_code, "clubs_data.json", [])
                    if clubs:
                        for club in clubs:
                            with st.expander(f"🎭 {club['name']}"):
                                st.write(f"**Patron:** {club['patron']}")
                                st.write(f"**Venue:** {club['venue']}")
                                st.write(f"**Schedule:** {club['schedule']}")
                                st.write(f"**Members:** {len(club.get('members', []))}")
                
                with mgmt_tabs[8]:  # Transport
                    st.subheader("Transport Management")
                    
                    with st.expander("➕ Add Route"):
                        with st.form("add_route"):
                            route_name = st.text_input("Route Name")
                            driver = st.text_input("Driver Name")
                            vehicle = st.text_input("Vehicle Registration")
                            departure = st.time_input("Departure Time")
                            capacity = st.number_input("Capacity", 1, 100, 40)
                            
                            if st.form_submit_button("Add Route"):
                                add_route(school_code, route_name, driver, vehicle, str(departure), capacity)
                                st.success("Route added!")
                                st.rerun()
                    
                    transport = load_school_data(school_code, "transport_data.json", [])
                    if transport:
                        df = pd.DataFrame(transport)
                        st.dataframe(df[['name', 'driver', 'vehicle', 'departure_time', 'capacity']])
                
                with mgmt_tabs[9]:  # Cafeteria
                    st.subheader("Cafeteria Management")
                    
                    with st.expander("➕ Add Menu Item"):
                        with st.form("add_menu"):
                            item_name = st.text_input("Item Name")
                            category = st.selectbox("Category", ["Breakfast", "Lunch", "Snacks", "Beverages"])
                            price = st.number_input("Price (KES)", 0.0, step=10.0)
                            quantity = st.number_input("Available Quantity", 1, 1000, 50)
                            day = st.selectbox("Day", ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "All Week"])
                            
                            if st.form_submit_button("Add Item"):
                                add_menu_item(school_code, item_name, category, price, quantity, day)
                                st.success("Menu item added!")
                                st.rerun()
                    
                    cafeteria = load_school_data(school_code, "cafeteria_data.json", [])
                    if cafeteria:
                        df = pd.DataFrame(cafeteria)
                        st.dataframe(df[['name', 'category', 'price', 'available', 'day']])
                
                with mgmt_tabs[10]:  # Health
                    st.subheader("Health Records")
                    
                    with st.expander("➕ Add Health Record"):
                        students = [u for u in users if u['role'] == 'student']
                        with st.form("add_health"):
                            student = st.selectbox("Select Student", [f"{s['fullname']} ({s['admission_number']})" for s in students])
                            record_type = st.selectbox("Record Type", ["Sick Visit", "Medication", "Vaccination", "Check-up", "Emergency"])
                            description = st.text_area("Description")
                            
                            if st.form_submit_button("Save Record"):
                                student_email = student.split('(')[1].rstrip(')')
                                add_health_record(
                                    school_code, student_email, record_type, description,
                                    datetime.now().strftime("%Y-%m-%d"), st.session_state.user['email']
                                )
                                st.success("Health record saved!")
                                st.rerun()
                    
                    health = load_school_data(school_code, "health_records.json", [])
                    if health:
                        df = pd.DataFrame(health)
                        st.dataframe(df[['student_email', 'type', 'date']])
                
                with mgmt_tabs[11]:  # Counseling
                    st.subheader("Counseling Records")
                    
                    with st.expander("➕ Add Counseling Session"):
                        students = [u for u in users if u['role'] == 'student']
                        with st.form("add_counseling"):
                            student = st.selectbox("Select Student", [f"{s['fullname']} ({s['admission_number']})" for s in students])
                            session_type = st.selectbox("Session Type", ["Academic", "Career", "Personal", "Group", "Family"])
                            notes = st.text_area("Session Notes")
                            
                            if st.form_submit_button("Save Session"):
                                student_email = student.split('(')[1].rstrip(')')
                                add_counseling_session(
                                    school_code, student_email, st.session_state.user['fullname'],
                                    session_type, notes, datetime.now().strftime("%Y-%m-%d")
                                )
                                st.success("Counseling session saved!")
                                st.rerun()
                    
                    counseling = load_school_data(school_code, "counseling_records.json", [])
                    if counseling:
                        df = pd.DataFrame(counseling)
                        st.dataframe(df[['student_email', 'counselor', 'type', 'date']])
                
                with mgmt_tabs[12]:  # Alumni
                    st.subheader("Alumni Management")
                    
                    with st.expander("➕ Add Alumni"):
                        with st.form("add_alumni"):
                            name = st.text_input("Full Name")
                            year = st.number_input("Graduation Year", 1950, datetime.now().year)
                            occupation = st.text_input("Current Occupation")
                            contact = st.text_input("Email/Phone")
                            achievements = st.text_area("Achievements")
                            
                            if st.form_submit_button("Add Alumni"):
                                add_alumni(school_code, name, str(year), occupation, contact, achievements)
                                st.success("Alumni added!")
                                st.rerun()
                    
                    alumni = load_school_data(school_code, "alumni_data.json", [])
                    if alumni:
                        df = pd.DataFrame(alumni)
                        st.dataframe(df[['name', 'graduation_year', 'occupation']])
                
                with mgmt_tabs[13]:  # Staff
                    st.subheader("Staff Management")
                    
                    with st.expander("➕ Add Staff"):
                        with st.form("add_staff"):
                            name = st.text_input("Full Name")
                            position = st.text_input("Position")
                            department = st.selectbox("Department", ["Academic", "Administration", "Support", "Maintenance", "Security"])
                            employment_date = st.date_input("Employment Date")
                            contact = st.text_input("Contact")
                            qualifications = st.text_area("Qualifications")
                            
                            if st.form_submit_button("Add Staff"):
                                add_staff(
                                    school_code, name, position, department,
                                    employment_date.strftime("%Y-%m-%d"), contact, qualifications
                                )
                                st.success("Staff added!")
                                st.rerun()
                    
                    staff = load_school_data(school_code, "staff_data.json", [])
                    if staff:
                        df = pd.DataFrame(staff)
                        st.dataframe(df[['name', 'position', 'department', 'employment_date']])
                
                with mgmt_tabs[14]:  # Calendar
                    st.subheader("School Calendar")
                    
                    with st.expander("➕ Add Event"):
                        with st.form("add_event"):
                            title = st.text_input("Event Title")
                            event_date = st.date_input("Event Date")
                            event_type = st.selectbox("Event Type", ["Holiday", "Exam", "Meeting", "Sports Day", "Graduation", "Other"])
                            description = st.text_area("Description")
                            audience = st.multiselect("Audience", ["Students", "Teachers", "Parents", "Staff", "All"])
                            location = st.text_input("Location")
                            
                            if st.form_submit_button("Add Event"):
                                add_calendar_event(
                                    school_code, title, event_date.strftime("%Y-%m-%d"),
                                    event_type, description, audience, location
                                )
                                st.success("Event added!")
                                st.rerun()
                    
                    upcoming = get_upcoming_events(school_code)
                    if upcoming:
                        for event in upcoming:
                            st.markdown(f"""
                            <div class="golden-card">
                                <strong>{event['title']}</strong> - {event['date']}<br>
                                Type: {event['type']} | Location: {event['location']}<br>
                                {event['description']}
                            </div>
                            """, unsafe_allow_html=True)
                
                with mgmt_tabs[15]:  # Reports
                    st.subheader("Generate Reports")
                    
                    report_type = st.selectbox("Report Type", [
                        "Academic Performance", "Attendance Summary", "Financial Report",
                        "Discipline Report", "Library Usage", "Sports Participation",
                        "Alumni Directory", "Staff Directory", "Inventory Status"
                    ])
                    
                    if st.button("Generate Report"):
                        if report_type == "Academic Performance":
                            grades = load_school_data(school_code, "academic_records.json", [])
                            if grades:
                                df = pd.DataFrame(grades)
                                csv = df.to_csv(index=False)
                                st.download_button("Download CSV", csv, "academic_report.csv", "text/csv")
                                st.success("Report generated!")
                        
                        elif report_type == "Financial Report":
                            fees = load_school_data(school_code, "fees.json", [])
                            if fees:
                                df = pd.DataFrame(fees)
                                csv = df.to_csv(index=False)
                                st.download_button("Download CSV", csv, "financial_report.csv", "text/csv")
                                st.success("Report generated!")
                
                with mgmt_tabs[16]:  # Analytics
                    st.subheader("Performance Analytics")
                    
                    term = st.selectbox("Select Term", TERMS)
                    year = st.selectbox("Select Year", [str(y) for y in range(2020, datetime.now().year+1)])
                    
                    if st.button("Generate Heatmap"):
                        fig = generate_performance_heatmap(school_code, term, year)
                        if fig:
                            st.plotly_chart(fig)
                        else:
                            st.info("No data available for this term")
                    
                    trends = analyze_performance_trends(school_code)
                    if trends:
                        st.subheader("Performance Trends")
                        if trends["overall_trend"]:
                            df = pd.DataFrame(trends["overall_trend"])
                            fig = px.line(df, x='date', y='score', title="Overall Performance Trend")
                            st.plotly_chart(fig)
                
                with mgmt_tabs[17]:  # Predictions
                    st.subheader("Performance Predictions")
                    
                    students = [u for u in users if u['role'] == 'student']
                    student = st.selectbox("Select Student", [f"{s['fullname']} ({s['admission_number']})" for s in students])
                    
                    if st.button("Predict Performance"):
                        student_email = student.split('(')[1].rstrip(')')
                        predictions = predict_future_performance(school_code, student_email)
                        if predictions:
                            for subject, score in predictions.items():
                                st.metric(f"{subject} Predicted Score", f"{score}%")
                        else:
                            st.info("Insufficient data for prediction")
                
                with mgmt_tabs[18]:  # Inventory
                    st.subheader("Inventory Management")
                    
                    with st.expander("➕ Add Inventory Item"):
                        with st.form("add_inventory"):
                            item_name = st.text_input("Item Name")
                            category = st.selectbox("Category", ["Stationery", "Cleaning", "Maintenance", "Electronics", "Furniture", "Sports", "Other"])
                            quantity = st.number_input("Quantity", 0, 10000, 10)
                            unit = st.text_input("Unit (e.g., pcs, kg, liters)")
                            reorder_level = st.number_input("Reorder Level", 0, 1000, 5)
                            location = st.text_input("Storage Location")
                            
                            if st.form_submit_button("Add Item"):
                                add_inventory_item(school_code, item_name, category, quantity, unit, reorder_level, location)
                                st.success("Item added!")
                                st.rerun()
                    
                    inventory = load_school_data(school_code, "inventory.json", [])
                    if inventory:
                        df = pd.DataFrame(inventory)
                        st.dataframe(df[['name', 'category', 'quantity', 'unit', 'location']])
                        
                        low_stock = get_low_stock_items(school_code)
                        if low_stock:
                            st.warning(f"⚠️ {len(low_stock)} items are below reorder level!")
                
                with mgmt_tabs[19]:  # Budget
                    st.subheader("Budget Management")
                    
                    with st.expander("➕ Add Budget Item"):
                        with st.form("add_budget"):
                            category = st.selectbox("Category", ["Salaries", "Operations", "Maintenance", "Events", "Equipment", "Utilities", "Other"])
                            description = st.text_input("Description")
                            amount = st.number_input("Allocated Amount (KES)", 0.0, step=10000.0)
                            fiscal_year = st.text_input("Fiscal Year", f"{datetime.now().year}")
                            
                            if st.form_submit_button("Add Budget Item"):
                                add_budget_item(school_code, category, description, amount, fiscal_year)
                                st.success("Budget item added!")
                                st.rerun()
                    
                    budget = load_school_data(school_code, "budget.json", [])
                    if budget:
                        df = pd.DataFrame(budget)
                        st.dataframe(df[['category', 'description', 'allocated', 'spent', 'remaining']])
                        
                        total_allocated = sum(b['allocated'] for b in budget)
                        total_spent = sum(b['spent'] for b in budget)
                        
                        col1, col2, col3 = st.columns(3)
                        col1.metric("Total Allocated", f"KES {total_allocated:,.0f}")
                        col2.metric("Total Spent", f"KES {total_spent:,.0f}")
                        col3.metric("Remaining", f"KES {total_allocated - total_spent:,.0f}")
                
                with mgmt_tabs[20]:  # Settings
                    st.subheader("System Settings")
                    
                    if st.session_state.user['role'] == 'admin':
                        st.metric("Total Users", len(users))
                        st.metric("Students", len([u for u in users if u['role'] == 'student']))
                        st.metric("Teachers", len([u for u in users if u['role'] == 'teacher']))
                        st.metric("Guardians", len([u for u in users if u['role'] == 'guardian']))
                        
                        with st.form("update_school"):
                            school = st.session_state.current_school
                            name = st.text_input("School Name", school['name'])
                            motto = st.text_input("Motto", school.get('motto', ''))
                            city = st.text_input("City", school.get('city', ''))
                            state = st.text_input("State", school.get('state', ''))
                            
                            if st.form_submit_button("Update School Info"):
                                all_schools = load_all_schools()
                                all_schools[school_code]['name'] = name
                                all_schools[school_code]['motto'] = motto
                                all_schools[school_code]['city'] = city
                                all_schools[school_code]['state'] = state
                                save_all_schools(all_schools)
                                st.session_state.current_school = all_schools[school_code]
                                st.success("School information updated!")
                        
                        if st.button("Backup Data"):
                            # Create backup
                            backup_dir = DATA_DIR / "backups"
                            backup_dir.mkdir(exist_ok=True)
                            backup_file = backup_dir / f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
                            
                            with zipfile.ZipFile(backup_file, 'w') as zipf:
                                for file in DATA_DIR.glob(f"{school_code}_*.json"):
                                    zipf.write(file, file.name)
                            
                            st.success(f"Backup created: {backup_file.name}")
                            st.download_button(
                                "Download Backup",
                                open(backup_file, 'rb').read(),
                                backup_file.name,
                                "application/zip"
                            )
                    else:
                        st.warning("Admin access required for system settings.")
            else:
                st.warning("⚠️ You need admin or teacher access for School Management. Please login with appropriate credentials.")
                if st.button("Go to Login"):
                    st.session_state.main_section = 'community'
                    st.rerun()
        else:
            st.warning("⚠️ Please login first to access School Management.")
            if st.button("Go to Login"):
                st.session_state.main_section = 'community'
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    
    # ============ PERSONAL DASHBOARD SECTION ============
    elif st.session_state.main_section == 'personal':
        st.markdown('<div class="section-container section-personal">', unsafe_allow_html=True)
        st.markdown("""
        <div style="background: linear-gradient(135deg, #45B7D1, #6EC8E0); padding: 30px; border-radius: 20px; text-align: center; margin-bottom: 30px; box-shadow: 0 0 40px rgba(69, 183, 209, 0.5);">
            <h3 style="color: white; text-shadow: 2px 2px 4px rgba(0,0,0,0.3);">👤 Personal Dashboard</h3>
            <p style="color: white;">Your personal information, performance, and analytics - Requires login</p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.session_state.user and st.session_state.current_school:
            user = st.session_state.user
            school_code = st.session_state.current_school['code']
            users = load_school_data(school_code, "users.json", [])
            
            personal_tabs = st.tabs([
                "👤 Profile", "📊 My Performance", "📈 Class Analytics", "⭐ Reviews",
                "💰 Fees", "📋 Attendance", "⚠️ Discipline", "🏥 Health",
                "🧠 Counseling", "🏃 Sports", "🎭 Clubs", "🚌 Transport",
                "🍽️ Cafeteria", "📚 Library", "📅 Calendar", "🔮 Predictions"
            ])
            
            with personal_tabs[0]:  # Profile
                st.markdown("#### Personal Information")
                col1, col2 = st.columns([1, 2])
                
                with col1:
                    if user.get('profile_pic'):
                        st.image(user['profile_pic'], width=150)
                    else:
                        emoji = "👑" if user['role'] == 'admin' else "👨‍🏫" if user['role'] == 'teacher' else "👨‍🎓" if user['role'] == 'student' else "👪"
                        st.markdown(f"<h1 style='font-size: 5rem; text-align: center;'>{emoji}</h1>", unsafe_allow_html=True)
                    
                    pic = st.file_uploader("📸 Upload Photo", type=['png', 'jpg', 'jpeg'])
                    if pic:
                        img = Image.open(pic)
                        buffered = BytesIO()
                        img.save(buffered, format="PNG")
                        b64 = base64.b64encode(buffered.getvalue()).decode()
                        
                        for u in users:
                            if u['email'] == user['email']:
                                u['profile_pic'] = f"data:image/png;base64,{b64}"
                        save_school_data(school_code, "users.json", users)
                        user['profile_pic'] = f"data:image/png;base64,{b64}"
                        st.rerun()
                
                with col2:
                    with st.form("update_profile"):
                        fullname = st.text_input("Full Name", user['fullname'])
                        phone = st.text_input("Phone", user.get('phone', ''))
                        bio = st.text_area("Bio", user.get('bio', ''), height=100)
                        
                        if user['role'] == 'student':
                            st.info(f"🎫 Admission: **{user['admission_number']}**")
                            st.info(f"📚 Class: **{user['class']}**")
                            st.info(f"🌊 Stream: **{user['stream']}**")
                        elif user['role'] == 'guardian':
                            linked = user.get('linked_students', [])
                            st.info(f"👪 Linked Students: {', '.join(linked)}")
                        elif user['role'] == 'teacher':
                            allocations = get_teacher_allocations(school_code, user['fullname'])
                            if allocations:
                                st.info(f"📚 Teaching: {', '.join([a['subject'] for a in allocations])}")
                        
                        if st.form_submit_button("Update"):
                            for u in users:
                                if u['email'] == user['email']:
                                    u['fullname'] = fullname
                                    u['phone'] = phone
                                    u['bio'] = bio
                            save_school_data(school_code, "users.json", users)
                            user.update({'fullname': fullname, 'phone': phone, 'bio': bio})
                            st.success("Profile updated!")
                            st.rerun()
            
            with personal_tabs[1]:  # My Performance
                st.markdown("#### My Performance")
                
                if user['role'] == 'student':
                    grades = load_school_data(school_code, "academic_records.json", [])
                    attendance = load_school_data(school_code, "attendance.json", [])
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        perf = calculate_student_performance(grades, user['email'])
                        st.metric("Overall Average", f"{perf['average']}%")
                        st.metric("Total Points", perf['total_points'])
                        
                        rank_class = "performance-excellent" if perf['average'] >= 80 else \
                                     "performance-good" if perf['average'] >= 70 else \
                                     "performance-average" if perf['average'] >= 50 else \
                                     "performance-needs-improvement"
                        st.markdown(f"<div class='{rank_class}' style='padding:10px; text-align:center;'>{perf['rank']}</div>", unsafe_allow_html=True)
                        
                        if perf['subjects']:
                            df = pd.DataFrame(list(perf['subjects'].items()), columns=['Subject', 'Score'])
                            fig = px.bar(df, x='Subject', y='Score', title="Subject Performance", 
                                       color='Score', color_continuous_scale='RdYlGn')
                            st.plotly_chart(fig, use_container_width=True)
                    
                    with col2:
                        student_attendance = [a for a in attendance if a['student_email'] == user['email']]
                        if student_attendance:
                            present = len([a for a in student_attendance if a['status'] == 'Present'])
                            absent = len([a for a in student_attendance if a['status'] == 'Absent'])
                            late = len([a for a in student_attendance if a['status'] == 'Late'])
                            
                            att_df = pd.DataFrame({'Status': ['Present', 'Absent', 'Late'], 'Count': [present, absent, late]})
                            fig = px.pie(att_df, values='Count', names='Status', title="Attendance",
                                       color_discrete_sequence=['#28a745', '#dc3545', '#ffc107'])
                            st.plotly_chart(fig, use_container_width=True)
                            
                            rate = (present / len(student_attendance)) * 100 if student_attendance else 0
                            st.metric("Attendance Rate", f"{rate:.1f}%")
                        else:
                            st.info("No attendance records")
                    
                    # Performance Trend
                    st.subheader("Performance Trend")
                    if grades:
                        my_grades = [g for g in grades if g['student_email'] == user['email']]
                        if len(my_grades) >= 3:
                            df = pd.DataFrame(my_grades)
                            df['date'] = pd.to_datetime(df['date'])
                            df = df.sort_values('date')
                            fig = px.line(df, x='date', y='score', color='subject', title="Performance Over Time")
                            st.plotly_chart(fig)
                
                elif user['role'] == 'teacher':
                    st.subheader("Classes Taught")
                    allocations = get_teacher_allocations(school_code, user['fullname'])
                    if allocations:
                        for a in allocations:
                            st.info(f"📚 {a['subject']} - {a['assigned_class']} ({a['lessons_per_week']} lessons/week)")
                    
                    st.subheader("Reviews Given")
                    reviews = get_teacher_reviews(school_code, user['email'])
                    st.metric("Total Reviews", len(reviews))
                
                elif user['role'] == 'guardian':
                    for adm in user.get('linked_students', []):
                        student = next((u for u in users if u.get('admission_number') == adm), None)
                        if student:
                            st.subheader(f"👤 {student['fullname']}")
                            grades = load_school_data(school_code, "academic_records.json", [])
                            perf = calculate_student_performance(grades, student['email'])
                            st.metric("Average", f"{perf['average']}%")
                            
                            if perf['subjects']:
                                df = pd.DataFrame(list(perf['subjects'].items()), columns=['Subject', 'Score'])
                                fig = px.bar(df, x='Subject', y='Score', title=f"{student['fullname']}'s Performance")
                                st.plotly_chart(fig)
            
            with personal_tabs[2]:  # Class Analytics
                st.markdown("#### Class Performance Analytics")
                
                if user['role'] == 'student':
                    grades = load_school_data(school_code, "academic_records.json", [])
                    my_grades = [g for g in grades if g['student_email'] == user['email']]
                    
                    if my_grades:
                        my_class = user['class']
                        my_subjects = {g['subject'] for g in my_grades}
                        
                        for subject in my_subjects:
                            class_grades = [g['score'] for g in grades if g['class_name'] == my_class and g['subject'] == subject]
                            my_score = next(g['score'] for g in my_grades if g['subject'] == subject)
                            
                            if class_grades:
                                avg_class = sum(class_grades) / len(class_grades)
                                max_class = max(class_grades)
                                min_class = min(class_grades)
                                
                                col1, col2, col3, col4 = st.columns(4)
                                col1.metric(f"{subject}", f"{my_score}%")
                                col2.metric("Class Avg", f"{avg_class:.1f}%")
                                col3.metric("Class Max", f"{max_class:.1f}%")
                                col4.metric("Class Min", f"{min_class:.1f}%")
                                
                                # Position in class
                                sorted_grades = sorted(class_grades, reverse=True)
                                position = sorted_grades.index(my_score) + 1
                                st.progress(position/len(class_grades), text=f"Position: {position}/{len(class_grades)}")
                                st.divider()
                
                elif user['role'] == 'teacher':
                    st.subheader("Class Performance Overview")
                    allocations = get_teacher_allocations(school_code, user['fullname'])
                    
                    for alloc in allocations:
                        grades = load_school_data(school_code, "academic_records.json", [])
                        class_perf = [g for g in grades if g['class_name'] == alloc['assigned_class'] and g['subject'] == alloc['subject']]
                        
                        if class_perf:
                            scores = [g['score'] for g in class_perf]
                            df = pd.DataFrame({
                                'Student': [g['student_email'] for g in class_perf],
                                'Score': scores
                            }).sort_values('Score', ascending=False)
                            
                            st.write(f"**{alloc['assigned_class']} - {alloc['subject']}**")
                            st.dataframe(df)
                            
                            fig = px.histogram(df, x='Score', nbins=20, title=f"{alloc['subject']} Score Distribution")
                            st.plotly_chart(fig)
            
            with personal_tabs[3]:  # Reviews
                st.markdown("#### Reviews & Feedback")
                
                if user['role'] == 'student':
                    reviews = load_school_data(school_code, "teacher_reviews.json", [])
                    my_reviews = [r for r in reviews if r['student_email'] == user['email']]
                    
                    for r in my_reviews:
                        teacher = next((u for u in users if u['email'] == r['teacher_email']), None)
                        teacher_name = teacher['fullname'] if teacher else r['teacher_email']
                        st.markdown(f"""
                        <div class="golden-card">
                            <strong>From: {teacher_name}</strong><br>
                            ⭐ {'⭐' * r['rating']}{'☆' * (5-r['rating'])}<br>
                            📅 {r['date']}<br>
                            💬 {r['review_text']}
                        </div>
                        """, unsafe_allow_html=True)
                
                elif user['role'] == 'teacher':
                    with st.form("give_review"):
                        students = [u for u in users if u['role'] == 'student']
                        student = st.selectbox("Student", [f"{s['fullname']} ({s['admission_number']})" for s in students])
                        rating = st.slider("Rating", 1, 5, 3)
                        review = st.text_area("Review")
                        
                        if st.form_submit_button("Submit"):
                            student_email = student.split('(')[1].rstrip(')')
                            student_obj = next(s for s in students if s['email'] == student_email)
                            add_teacher_review(
                                school_code, user['email'], student_email, review, rating,
                                datetime.now().strftime("%Y-%m-%d"), student_obj['class']
                            )
                            st.success("Review submitted!")
                            st.rerun()
                    
                    feedback = load_school_data(school_code, "parent_feedback.json", [])
                    if feedback:
                        st.subheader("Parent Feedback")
                        for fb in feedback[-5:]:
                            parent = next((u for u in users if u['email'] == fb['guardian_email']), None)
                            parent_name = parent['fullname'] if parent else fb['guardian_email']
                            st.info(f"**{parent_name}**: {fb['feedback_text']} ({fb['date']})")
                
                elif user['role'] == 'guardian':
                    with st.form("give_feedback"):
                        students = [u for u in users if u.get('admission_number') in user.get('linked_students', [])]
                        student = st.selectbox("Select Student", [f"{s['fullname']} ({s['admission_number']})" for s in students])
                        feedback = st.text_area("Your Feedback")
                        
                        if st.form_submit_button("Submit"):
                            student_email = student.split('(')[1].rstrip(')')
                            add_parent_feedback(
                                school_code, user['email'], student_email, feedback,
                                datetime.now().strftime("%Y-%m-%d")
                            )
                            st.success("Feedback submitted!")
                            st.rerun()
            
            with personal_tabs[4]:  # Fees
                st.markdown("#### Fee Status")
                
                if user['role'] == 'student':
                    fees = load_school_data(school_code, "fees.json", [])
                    student_fees = [f for f in fees if f['student_email'] == user['email']]
                    
                    if student_fees:
                        total_paid = sum(f['amount'] for f in student_fees if f['status'] == 'Paid')
                        total_pending = sum(f['amount'] for f in student_fees if f['status'] in ['Pending', 'Overdue'])
                        balance = calculate_fee_balance(school_code, user['email'])
                        
                        col1, col2, col3 = st.columns(3)
                        col1.metric("Total Paid", f"KES {total_paid:,.0f}")
                        col2.metric("Pending", f"KES {total_pending:,.0f}")
                        col3.metric("Balance", f"KES {balance:,.0f}", delta_color="inverse")
                        
                        df = pd.DataFrame(student_fees)
                        st.dataframe(df[['date', 'type', 'amount', 'status', 'receipt_no']])
                    else:
                        st.info("No fee records found")
                
                elif user['role'] == 'guardian':
                    for adm in user.get('linked_students', []):
                        student = next((u for u in users if u.get('admission_number') == adm), None)
                        if student:
                            st.subheader(f"{student['fullname']}")
                            fees = load_school_data(school_code, "fees.json", [])
                            student_fees = [f for f in fees if f['student_email'] == student['email']]
                            
                            if student_fees:
                                total_paid = sum(f['amount'] for f in student_fees if f['status'] == 'Paid')
                                total_pending = sum(f['amount'] for f in student_fees if f['status'] in ['Pending', 'Overdue'])
                                
                                col1, col2 = st.columns(2)
                                col1.metric("Paid", f"KES {total_paid:,.0f}")
                                col2.metric("Pending", f"KES {total_pending:,.0f}")
            
            with personal_tabs[5]:  # Attendance
                st.markdown("#### Attendance Records")
                
                if user['role'] == 'student':
                    attendance = load_school_data(school_code, "attendance.json", [])
                    student_attendance = [a for a in attendance if a['student_email'] == user['email']]
                    
                    if student_attendance:
                        df = pd.DataFrame(student_attendance)
                        st.dataframe(df[['date', 'status', 'remarks']])
                        
                        # Monthly chart
                        df['month'] = pd.to_datetime(df['date']).dt.month
                        monthly = df.groupby(['month', 'status']).size().reset_index(name='count')
                        fig = px.bar(monthly, x='month', y='count', color='status', title="Monthly Attendance")
                        st.plotly_chart(fig)
                    else:
                        st.info("No attendance records found")
            
            with personal_tabs[6]:  # Discipline
                st.markdown("#### Discipline History")
                
                if user['role'] == 'student':
                    discipline = get_student_discipline_history(school_code, user['email'])
                    if discipline:
                        df = pd.DataFrame(discipline)
                        st.dataframe(df[['date', 'incident', 'action_taken', 'severity']])
                    else:
                        st.success("No discipline records - Good standing!")
            
            with personal_tabs[7]:  # Health
                st.markdown("#### Health Records")
                
                if user['role'] == 'student':
                    health = get_student_health_records(school_code, user['email'])
                    if health:
                        df = pd.DataFrame(health)
                        st.dataframe(df[['date', 'type', 'description']])
                    else:
                        st.info("No health records found")
            
            with personal_tabs[8]:  # Counseling
                st.markdown("#### Counseling Sessions")
                
                if user['role'] == 'student':
                    counseling = get_student_counseling(school_code, user['email'])
                    if counseling:
                        df = pd.DataFrame(counseling)
                        st.dataframe(df[['date', 'counselor', 'type', 'notes']])
                    else:
                        st.info("No counseling records found")
            
            with personal_tabs[9]:  # Sports
                st.markdown("#### Sports Participation")
                
                if user['role'] == 'student':
                    sports = get_student_sports(school_code, user['email'])
                    if sports:
                        for sport in sports:
                            st.markdown(f"""
                            <div class="golden-card">
                                <strong>{sport['sport']}</strong><br>
                                Coach: {sport['coach']} | Position: {sport.get('position', 'Member')}<br>
                                Joined: {sport['joined']}
                            </div>
                            """, unsafe_allow_html=True)
                    else:
                        st.info("Not participating in any sports")
            
            with personal_tabs[10]:  # Clubs
                st.markdown("#### Club Memberships")
                
                if user['role'] == 'student':
                    clubs = get_student_clubs(school_code, user['email'])
                    if clubs:
                        for club in clubs:
                            st.markdown(f"""
                            <div class="golden-card">
                                <strong>{club['club']}</strong><br>
                                Patron: {club['patron']} | Role: {club.get('role', 'Member')}<br>
                                Joined: {club['joined']}
                            </div>
                            """, unsafe_allow_html=True)
                    else:
                        st.info("Not a member of any clubs")
            
            with personal_tabs[11]:  # Transport
                st.markdown("#### Transport Details")
                
                if user['role'] == 'student':
                    transport = get_student_transport(school_code, user['email'])
                    if transport:
                        st.markdown(f"""
                        <div class="golden-card">
                            <strong>Route: {transport['route']}</strong><br>
                            Driver: {transport['driver']} | Vehicle: {transport['vehicle']}<br>
                            Departure: {transport['departure']} | Pickup: {transport['pickup']}
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.info("No transport assigned")
            
            with personal_tabs[12]:  # Cafeteria
                st.markdown("#### Cafeteria Purchases")
                
                if user['role'] == 'student':
                    purchases = load_school_data(school_code, "cafeteria_purchases.json", [])
                    my_purchases = [p for p in purchases if p['student_email'] == user['email']]
                    
                    if my_purchases:
                        df = pd.DataFrame(my_purchases)
                        st.dataframe(df[['purchase_date', 'total_amount', 'payment_method']])
                        
                        total_spent = sum(p['total_amount'] for p in my_purchases)
                        st.metric("Total Spent", f"KES {total_spent:,.0f}")
                    else:
                        st.info("No cafeteria purchases found")
            
            with personal_tabs[13]:  # Library
                st.markdown("#### Library Books")
                
                books = load_school_data(school_code, "library_books.json", [])
                my_books = []
                for book in books:
                    for issue in book.get('issued_to', []):
                        if issue['user_email'] == user['email'] and not issue['returned']:
                            my_books.append({
                                "title": book['title'],
                                "author": book['author'],
                                "issue_date": issue['issue_date'],
                                "due_date": issue['due_date']
                            })
                
                if my_books:
                    df = pd.DataFrame(my_books)
                    st.dataframe(df)
                    
                    # Check overdue
                    today = datetime.now().date()
                    overdue = [b for b in my_books if datetime.strptime(b['due_date'], "%Y-%m-%d").date() < today]
                    if overdue:
                        st.warning(f"⚠️ You have {len(overdue)} overdue books!")
                else:
                    st.info("No books currently issued")
            
            with personal_tabs[14]:  # Calendar
                st.markdown("#### Upcoming Events")
                
                events = get_upcoming_events(school_code)
                if events:
                    for event in events[:10]:
                        st.markdown(f"""
                        <div class="golden-card">
                            <strong>{event['title']}</strong> - {event['date']}<br>
                            Type: {event['type']} | Location: {event['location']}
                        </div>
                        """, unsafe_allow_html=True)
                else:
                    st.info("No upcoming events")
            
            with personal_tabs[15]:  # Predictions
                st.markdown("#### Performance Predictions")
                
                if user['role'] == 'student':
                    predictions = predict_future_performance(school_code, user['email'])
                    if predictions:
                        for subject, score in predictions.items():
                            st.metric(f"{subject} - Next Term Prediction", f"{score}%")
                        
                        st.info("Based on your performance trend analysis")
                    else:
                        st.info("Insufficient data for predictions. Continue taking exams to get predictions!")
        else:
            st.warning("⚠️ Please login first to access Personal Dashboard.")
            if st.button("Go to Login"):
                st.session_state.main_section = 'community'
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

# ----- DASHBOARD (for logged in users with sidebar) -----
elif st.session_state.page == 'dashboard' and st.session_state.current_school and st.session_state.user:
    school = st.session_state.current_school
    user = st.session_state.user
    school_code = school['code']
    
    users = load_school_data(school_code, "users.json", [])
    classes = load_school_data(school_code, "classes.json", [])
    announcements = load_school_data(school_code, "announcements.json", [])
    assignments = load_school_data(school_code, "assignments.json", [])
    
    unread_count = get_unread_count(user['email'], school_code)
    pending_friend_count = len(get_pending_requests(school_code, user['email']))
    
    # ============ GOLDEN SIDEBAR ============
    with st.sidebar:
        st.markdown(f"""
        <div class="school-header">
            <h2>{school['name']}</h2>
            <div class="school-code">
                <code>{school['code']}</code>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown('<div class="profile-card">', unsafe_allow_html=True)
        if user.get('profile_pic'):
            st.image(user['profile_pic'], width=50)
        else:
            emoji = "👑" if user['role'] == 'admin' else "👨‍🏫" if user['role'] == 'teacher' else "👨‍🎓" if user['role'] == 'student' else "👪" if user['role'] == 'guardian' else "👤"
            st.markdown(f"<h1 style='font-size: 2rem; margin: 0;'>{emoji}</h1>", unsafe_allow_html=True)
        
        role_display = user['role'].upper()
        st.markdown(f"""
        <div style="color: #FFD700; flex: 1;">
            <strong>{user['fullname']}</strong><br>
            <span style="background: rgba(0,0,0,0.3); color: #FFD700; padding: 2px 8px; border-radius: 12px; font-size: 0.7rem;">{role_display}</span>
        </div>
        """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
        
        st.divider()
        
        # Navigation options based on role
        if user['role'] == 'admin':
            options = ["Dashboard", "Announcements", "Classes", "Groups", "Teachers", "Students", 
                      "Guardians", "Assignments", "Community", f"Chat 💬{f' ({unread_count})' if unread_count>0 else ''}", 
                      f"Group Chats 👥", f"Friends 🤝{f' ({pending_friend_count})' if pending_friend_count>0 else ''}", 
                      "School Management", "Personal Dashboard", "Profile"]
        elif user['role'] == 'teacher':
            options = ["Dashboard", "Announcements", "My Classes", "Groups", "Assignments", "Community", 
                      f"Chat 💬{f' ({unread_count})' if unread_count>0 else ''}", f"Group Chats 👥", 
                      f"Friends 🤝{f' ({pending_friend_count})' if pending_friend_count>0 else ''}", 
                      "School Management", "Personal Dashboard", "Profile"]
        elif user['role'] == 'student':
            options = ["Dashboard", "Announcements", "Browse Classes", "My Classes", "Groups", "Assignments", 
                      "Community", f"Chat 💬{f' ({unread_count})' if unread_count>0 else ''}", 
                      f"Group Chats 👥", f"Friends 🤝{f' ({pending_friend_count})' if pending_friend_count>0 else ''}", 
                      "Personal Dashboard", "Profile"]
        elif user['role'] == 'guardian':
            options = ["Dashboard", "Announcements", "My Student", "Assignments", "Community", 
                      f"Chat 💬{f' ({unread_count})' if unread_count>0 else ''}", f"Group Chats 👥", 
                      f"Friends 🤝{f' ({pending_friend_count})' if pending_friend_count>0 else ''}", 
                      "Personal Dashboard", "Profile"]
        else:  # visitor
            options = ["Dashboard", "Announcements", "Community", "Profile"]
        
        if st.session_state.menu_index >= len(options):
            st.session_state.menu_index = 0
            
        menu = st.radio("Navigation", options, index=st.session_state.menu_index, label_visibility="collapsed")
        st.session_state.menu_index = options.index(menu)
        
        st.divider()
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🏠 Home", use_container_width=True):
                st.session_state.page = 'welcome'
                st.rerun()
        with col2:
            if st.button("🚪 Logout", use_container_width=True):
                st.session_state.user = None
                st.session_state.current_school = None
                st.session_state.page = 'welcome'
                st.rerun()
    
    # ============ MAIN CONTENT ============
    if menu == "Dashboard":
        st.markdown(f"<h2 style='text-align: center;'>Welcome, {user['fullname']}!</h2>", unsafe_allow_html=True)
        
        if user['role'] == 'admin':
            col1, col2, col3, col4, col5, col6 = st.columns(6)
            col1.metric("Students", school['stats'].get('students', 0))
            col2.metric("Teachers", school['stats'].get('teachers', 0))
            col3.metric("Guardians", school['stats'].get('guardians', 0))
            col4.metric("Classes", school['stats'].get('classes', 0))
            col5.metric("Alumni", school['stats'].get('alumni', 0))
            col6.metric("Staff", len([u for u in users if u['role'] == 'staff']))
            
            st.subheader("📋 Recent Activity")
            recent_announcements = announcements[-3:] if announcements else []
            for ann in recent_announcements:
                st.info(f"📢 {ann['title']} - {ann['date'][:16]}")
        
        elif user['role'] == 'teacher':
            my_classes = [c for c in classes if c.get('teacher') == user['email']]
            allocations = get_teacher_allocations(school_code, user['fullname'])
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("My Classes", len(my_classes))
            with col2:
                st.metric("Subjects", len(allocations))
            with col3:
                my_assignments = len([a for a in assignments if a.get('created_by') == user['email']])
                st.metric("Assignments", my_assignments)
        
        elif user['role'] == 'student':
            my_classes = [c for c in classes if user['email'] in c.get('students', [])]
            my_groups = len(load_school_data(school_code, "groups.json", []))
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("My Classes", len(my_classes))
            with col2:
                st.metric("Admission", user['admission_number'][:10] + "...")
            with col3:
                st.metric("Class", user['class'].split('(')[0])
            with col4:
                st.metric("Stream", user['stream'])
            
            # Show fee balance
            balance = calculate_fee_balance(school_code, user['email'])
            if balance > 0:
                st.warning(f"⚠️ Fee Balance: KES {balance:,.0f}")
        
        elif user['role'] == 'guardian':
            linked = user.get('linked_students', [])
            st.info(f"👪 Linked to {len(linked)} student(s)")
            for adm in linked:
                student = next((u for u in users if u.get('admission_number') == adm), None)
                if student:
                    st.write(f"**{student['fullname']}** - {student['class']} ({student['stream']})")
        
        elif user['role'] == 'visitor':
            st.info("You are viewing as a guest. Login to access more features.")
    
    elif menu == "Announcements":
        st.markdown("<h2 style='text-align: center;'>📢 Announcements</h2>", unsafe_allow_html=True)
        if announcements:
            for a in announcements[-10:]:
                st.markdown(f"""
                <div class="golden-card">
                    <h4>{a['title']}</h4>
                    <p>{a['content']}</p>
                    <small>By {a['author']} • {a['date'][:16]}</small>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("No announcements")
    
    elif menu == "Profile":
        st.markdown("<h2 style='text-align: center;'>👤 Profile</h2>", unsafe_allow_html=True)
        col1, col2 = st.columns([1, 2])
        with col1:
            if user.get('profile_pic'):
                st.image(user['profile_pic'], width=150)
        with col2:
            st.write(f"**Name:** {user['fullname']}")
            st.write(f"**Email:** {user['email']}")
            st.write(f"**Role:** {user['role'].title()}")
            if user['role'] == 'student':
                st.write(f"**Admission:** {user['admission_number']}")
                st.write(f"**Class:** {user['class']}")
                st.write(f"**Stream:** {user['stream']}")
            elif user['role'] == 'guardian':
                st.write(f"**Linked Students:** {', '.join(user.get('linked_students', []))}")
            elif user['role'] == 'teacher':
                st.write(f"**Teacher Code:** {user.get('teacher_code', 'N/A')}")
    
    elif menu == "School Management":
        if user['role'] in ['admin', 'teacher']:
            st.markdown("<h2 style='text-align: center;'>📊 School Management</h2>", unsafe_allow_html=True)
            st.info("Redirecting to School Management section...")
            st.session_state.main_section = 'management'
            st.rerun()
        else:
            st.warning("Access denied. Admin/Teacher only.")
    
    elif menu == "Personal Dashboard":
        st.markdown("<h2 style='text-align: center;'>👤 Personal Dashboard</h2>", unsafe_allow_html=True)
        st.info("Redirecting to Personal Dashboard...")
        st.session_state.main_section = 'personal'
        st.rerun()
    
    elif menu == "Community":
        st.markdown("<h2 style='text-align: center;'>🌍 Community</h2>", unsafe_allow_html=True)
        st.info("Community features - Chat, friends, and groups will be available here")
        # Full community features implementation would go here
    
    elif menu.startswith("Chat"):
        st.markdown("<h2 style='text-align: center;'>💬 Chat</h2>", unsafe_allow_html=True)
        st.info("Chat feature coming soon!")
    
    elif menu == "Group Chats 👥":
        st.markdown("<h2 style='text-align: center;'>👥 Group Chats</h2>", unsafe_allow_html=True)
        st.info("Group chat feature coming soon!")
    
    elif menu.startswith("Friends"):
        st.markdown("<h2 style='text-align: center;'>🤝 Friends</h2>", unsafe_allow_html=True)
        st.info("Friends feature coming soon!")

else:
    st.error("Something went wrong. Please restart.")
    if st.button("Restart"):
        st.session_state.page = 'welcome'
        st.rerun()
